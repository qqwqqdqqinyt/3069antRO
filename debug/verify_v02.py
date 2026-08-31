# -*- coding: utf-8 -*-
"""v0.2 调参表结构校验：sheet 顺序 / 常量引用完整性 / 公式可解析性"""
import re, sys
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

FN = "Balance_星序防线_v0.2.xlsx"
wb = load_workbook(FN)
print("sheets:", wb.sheetnames)

# 1) 收集常量表 (01) 的 code -> 值/公式
ws1 = wb["01_核心常量"]
CONST = {}
for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
    a, d = row[0].value, row[3].value
    if isinstance(a, str) and re.fullmatch(r"[A-Z][A-Z0-9_]{1,24}", a.strip()):
        CONST[a.strip()] = d
print("常量个数:", len(CONST))
missing_val = [k for k, v in CONST.items() if v is None]
print("常量值为空的:", missing_val)

# 2) 全表扫描公式里的跨表引用是否都存在
bad = []
CELL = re.compile(r"'([^']+)'!\$?([A-Z]{1,3})\$?(\d+)")
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                for sheet, col, rowno in CELL.findall(v):
                    if sheet not in wb.sheetnames:
                        bad.append((ws.title, c.coordinate, "未知表 " + sheet))
                    else:
                        t = wb[sheet].cell(row=int(rowno), column=column_index_from_string(col))
                        if t.value is None:
                            bad.append((ws.title, c.coordinate, f"空引用 {sheet}!{col}{rowno}"))
print("坏引用数:", len(bad))
for b in bad[:20]:
    print("  ", b)

# 3) 统计公式数量
nf = sum(1 for ws in wb.worksheets for row in ws.iter_rows() for c in row
         if isinstance(c.value, str) and c.value.startswith("="))
print("公式单元格总数:", nf)
