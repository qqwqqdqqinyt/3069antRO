# -*- coding: utf-8 -*-
"""
v0.2 调参表公式求值器（离线校验用）
把 Excel 公式翻译成 Python 表达式并递归求值，用来在不开 Excel 的情况下
验证 572 个公式算得对不对。支持的语法子集：+ - * / ^ ( ) IF MAX MIN ROUND SUM LOG 比较运算。
"""
import math, re, sys
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as ci, get_column_letter as gl

FN = "Balance_星序防线_v0.2.xlsx"
wb = load_workbook(FN)
RAW = {}
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            RAW[(ws.title, c.coordinate)] = c.value

CACHE, STACK = {}, []


def rng(sheet, a, b):
    r1, c1 = int(re.match(r"[A-Z]+(\d+)", a).group(1)), ci(re.match(r"([A-Z]+)", a).group(1))
    r2, c2 = int(re.match(r"[A-Z]+(\d+)", b).group(1)), ci(re.match(r"([A-Z]+)", b).group(1))
    out = []
    for rr in range(min(r1, r2), max(r1, r2) + 1):
        for cc in range(min(c1, c2), max(c1, c2) + 1):
            v = val(sheet, f"{gl(cc)}{rr}")
            if isinstance(v, (int, float)):
                out.append(v)
    return out


def val(sheet, coord):
    key = (sheet, coord)
    if key in CACHE:
        return CACHE[key]
    if key in STACK:
        raise RuntimeError(f"循环引用 {sheet}!{coord} chain={STACK[-4:]}")
    raw = RAW.get(key)
    if raw is None:
        return 0
    if not (isinstance(raw, str) and raw.startswith("=")):
        CACHE[key] = raw
        return raw
    STACK.append(key)
    try:
        v = eval(translate(raw, sheet), ENV)
        CACHE[key] = v
        return v
    except Exception as e:
        CACHE[key] = f"#ERR({e})"
        return CACHE[key]
    finally:
        STACK.pop()


def _R(sheet, coord):
    return val(sheet, coord)


def _if(c, a, b):
    return a if c else b


def _log(x, base=10):
    x = float(x)
    if x <= 0:
        raise ValueError("LOG domain")
    return math.log(x, base)


def translate(f, sheet):
    s = f[1:]
    # 1) 跨表区域  'Sheet'!$C$5:$C$16
    s = re.sub(r"'([^']+)'!\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)",
               lambda m: f'_rng("{m.group(1)}","{m.group(2)}{m.group(3)}","{m.group(4)}{m.group(5)}")', s)
    # 2) 本表区域  C5:C16
    s = re.sub(r"(?<![A-Za-z0-9_!\"])\$?([A-Z]{1,3})\$?(\d+):\$?([A-Z]{1,3})\$?(\d+)",
               lambda m: f'_rng("{sheet}","{m.group(1)}{m.group(2)}","{m.group(3)}{m.group(4)}")', s)
    # 3) 跨表单格
    s = re.sub(r"'([^']+)'!\$?([A-Z]{1,3})\$?(\d+)",
               lambda m: f'_R("{m.group(1)}","{m.group(2)}{m.group(3)}")', s)
    # 4) 本表单格
    s = re.sub(r"(?<![A-Za-z0-9_!\"])\$?([A-Z]{1,3})\$?(\d+)(?!\()",
               lambda m: f'_R("{sheet}","{m.group(1)}{m.group(2)}")', s)
    s = re.sub(r"\bSUM\(", "_sum(", s)
    s = re.sub(r"\bIF\(", "_if(", s)
    s = re.sub(r"\bMIN\(", "min(", s)
    s = re.sub(r"\bMAX\(", "max(", s)
    s = re.sub(r"\bROUND\(", "round(", s)
    s = re.sub(r"\bLOG\(", "_log(", s)
    s = s.replace(">=", "@GE@").replace("<=", "@LE@").replace("<>", "@NE@")
    s = s.replace("=", "==")
    s = s.replace("@GE@", ">=").replace("@LE@", "<=").replace("@NE@", "!=")
    s = s.replace("^", "**")
    return s


def _rng_impl(sheet, a, b):
    return rng(sheet, a, b)


ENV = {"_R": _R, "_if": _if, "_log": _log, "min": min, "max": max, "round": round,
       "_rng": _rng_impl, "_sum": lambda x: sum(x)}


def V(sheet, coord):
    return val(sheet, coord)


if __name__ == "__main__":
    # 全表求值，收集错误
    errs = []
    for (sh, co), raw in RAW.items():
        if isinstance(raw, str) and raw.startswith("="):
            v = val(sh, co)
            if isinstance(v, str) and v.startswith("#ERR"):
                errs.append((sh, co, raw[:60], v))
    print("公式求值错误数:", len(errs))
    for e in errs[:15]:
        print("  ", e)
