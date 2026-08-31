# -*- coding: utf-8 -*-
"""
生成《星序：元素防线》调参总表 v0.2 —— 蒙特卡洛校准版
相对 v0.1 的核心变更：
  M1  CV / 充能归一化（CV' = v/E，充能' = k x log2(v/E) x Jensen修正）
  M2  生成值弱分层（x1.4/档）+ 棋盘跨关保留
  M3  输出占比翻转（附魔 60% : 植物 40%，EP 60->180，植物 DPS 12->7）
  M4  元素 AoE 总量封顶 EP x 2.5；删除连锁倍率
  M5  术语去重：核心(HP)->星枢、核心(货币)->晶核
新增 10_波次预算表（关 1 五波配平）
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---------- 样式 ----------
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄 = 可调输入
CALC_FILL  = PatternFill("solid", fgColor="EAF3DE")   # 绿 = 公式计算
MEAS_FILL  = PatternFill("solid", fgColor="DDEBF7")   # 蓝 = 蒙特卡洛实测值
HEAD_FILL  = PatternFill("solid", fgColor="4472C4")
HEAD_FONT  = Font(color="FFFFFF", bold=True, size=10)
SEC_FONT   = Font(bold=True, size=11, color="1F3864")
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
NOTE_FONT  = Font(size=9, color="808080", italic=True)
WARN_FONT  = Font(size=9, color="C00000", italic=True, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def head(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

def sec(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = SEC_FONT
    return row + 1

def note(ws, row, text, warn=False):
    c = ws.cell(row=row, column=1, value=text)
    c.font = WARN_FONT if warn else NOTE_FONT
    return row + 1

# ============================================================
# 00 说明
# ============================================================
ws = wb.active
ws.title = "00_说明"
ws.column_dimensions["A"].width = 24
ws.column_dimensions["B"].width = 112
ws["A1"] = "星序：元素防线 — 调参总表 v0.2（蒙特卡洛校准版）"
ws["A1"].font = TITLE_FONT
r = 3
r = note(ws, r, "本表所有经济公式已按 v0.2 的归一化方案重建。v0.1 的 xlsx 不可直接用于开发 —— 它的 CV 未归一化，跨档位膨胀 17 倍。", warn=True)
r += 1
r = sec(ws, r, "使用方式")
for a, b in [
    ("黄色单元格", "可调输入参数。改动后全表公式自动重算。"),
    ("绿色单元格", "公式计算结果。请勿手动覆盖，否则断链。"),
    ("蓝色单元格", "蒙特卡洛实测值（300–400 局 x 5 档 x 3–4 种玩家策略）。有模拟脚本可复现，改动前请先看脚本注释。"),
    ("灰色小字", "设计理由 / 调参方向说明。"),
    ("无底色", "静态参考数据（文案表、卡池清单等）。"),
]:
    ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=10)
    ws.cell(row=r, column=2, value=b).font = Font(size=10)
    r += 1
r += 1
r = sec(ws, r, "工作表导航")
for a, b in [
    ("01_核心常量", "全局旋钮。所有其他表的输入源头，优先在这里调。"),
    ("02_合成与充能", "★ 已归一化。方块值 -> 充能' / CV' / 各货币 / 附魔伤害，带档位选择器。"),
    ("03_波次节奏模型", "★ 已归一化。每波步数、合成次数、CV'/波、小附魔/波 —— 跨档位应基本恒定。"),
    ("04_关卡与敌人", "★ 换用新的技术模型（大数字达成率驱动）。1–15 关缩放与真实死亡关卡判定。"),
    ("05_植物", "★ DPS 已下调至辅助定位。含元素效果表（火的 AoE 已封顶）。"),
    ("06_卡牌", "★ 新增「战力点」列 —— 所有卡牌折算到同一把尺子，含超模/废卡处置建议。"),
    ("07_风险决策EV", "继续 / 收工的期望值计算器与保底比例敏感性分析。"),
    ("08_养成与花园", "星尘养成树成本模型与花园产出模型。"),
    ("09_验收清单", "平衡目标与「broken」判定标准 —— playtest 前先读这张。"),
    ("10_波次预算表", "★ 新增。关 1 五波的完整输出 / HP 配平，是敌人基准数值的来源。"),
]:
    ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=10)
    ws.cell(row=r, column=2, value=b).font = Font(size=10)
    r += 1
r += 1
r = sec(ws, r, "v0.2 五项结构性修正（详见 GDD_星序防线_v0.2.md）")
for a, b in [
    ("M1 归一化", "CV' = v / E[生成值]；充能' = k x log2(v/E) x 0.78(Jensen 修正)。原 CV 跨档膨胀 17 倍，归一化后波动 <2%。"),
    ("M2 弱分层 + 跨关保留", "分层增益 2.2x/档 -> 1.4x/档（强分层会让关 5 后技术倍差压到 1.0x）。棋盘跨关不重置。"),
    ("M3 输出占比翻转", "附魔 60% : 植物 40%（v0.1 实算是 21% : 79%，违反支柱 P2「合成即战斗」）。"),
    ("M4 AoE 封顶 / 删连锁", "元素总伤害池封顶 EP x 2.5 均分；连锁倍率会奖励乱滑，已删除。"),
    ("M5 术语去重", "核心(HP) -> 星枢；核心(货币) -> 晶核。"),
]:
    ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=10)
    c = ws.cell(row=r, column=2, value=b); c.font = Font(size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
r = note(ws, r, "模拟脚本（纯 stdlib，python sim_xxx.py <每档局数>）：sim_board_v02.py / sim_board_v02b.py / sim_chain_v02c.py / sim_tier_v02d.py")
r = note(ws, r, "配套文档：GDD_星序防线_v0.2.md（增量文档，未提及部分沿用 v0.1）")

# ============================================================
# 01 核心常量
# ============================================================
ws = wb.create_sheet("01_核心常量")
ws["A1"] = "核心常量 —— 全局旋钮（黄底可调 / 蓝底为蒙特卡洛实测值，改这里会驱动全表）"
ws["A1"].font = TITLE_FONT
head(ws, 3, ["分组", "参数名", "代号", "当前值", "建议范围", "调参说明"], [14, 28, 15, 12, 16, 66])

CONST = {}
r = 4
def const(group, name, code, val, rng, why, meas=False):
    global r
    ws.cell(row=r, column=1, value=group).border = BOX
    ws.cell(row=r, column=2, value=name).border = BOX
    ws.cell(row=r, column=3, value=code).border = BOX
    c = ws.cell(row=r, column=4, value=val)
    c.fill = MEAS_FILL if meas else INPUT_FILL
    c.border = BOX; c.font = Font(bold=True)
    ws.cell(row=r, column=5, value=rng).border = BOX
    c = ws.cell(row=r, column=6, value=why); c.border = BOX
    c.alignment = Alignment(wrap_text=True, vertical="top")
    CONST[code] = f"'01_核心常量'!$D${r}"
    r += 1

const("棋盘", "棋盘尺寸", "BOARD_N", 5, "5（不可改）", "★ 实测：4x4 在生成值分层后关 3+ 卡死率 88–92%，5x5 卡死率 <1.5%。4x4 会直接毁掉游戏。", meas=True)
const("棋盘", "每次移动生成方块数", "SPAWN_N", 1, "1 – 2", "保持 1。卡牌「丰饶」可临时 +1（但该卡已判定破坏性超模，见 06 表）。")
const("棋盘", "初始方块数", "INIT_N", 2, "2 – 4", "留出前期铺垫空间。")
const("棋盘", "步数上限", "STEP_MAX", 5, "3 – 10", "攒满 5 步可打一套爆发连招，是「攒劲儿」的爽点来源。")
const("棋盘", "步数回复间隔(秒)", "STEP_REGEN", 1.5, "1.0 – 2.5", "★ 手感主旋钮。直接决定每波步数预算。")
const("棋盘", "波次开局赠送步数", "STEP_GIFT", 2, "0 – 4", "缓解波次切换空窗。")
const("棋盘", "每步平均合成次数", "MERGE_RATE", 0.90, "0.85 – 0.95", "★★ 实测 0.89–0.94，跨策略差异 <6%。v0.1 假设的 1.2 是错的。这是 2048 的结构性常数，不要把技术水平建在这个量上。", meas=True)
const("棋盘", "平均合成值 / E[生成]", "AVG_MERGE_COEF", 5.47, "5.0 – 6.0", "★★ 实测：平均合成值 ≈ 5.47 x E[生成值]，跨全部 5 档稳定。所以归一化 CV'/波 = 合成次数 x 5.47，恒定。", meas=True)
const("棋盘", "充能 Jensen 修正", "LOG_CORR", 0.78, "0.70 – 0.85", "★★ 实测修正项。E[log2(v/E)] < log2(E[v]/E)，因为 log 是凹函数、合成值分布右偏。直接用 log2(均值) 会高估充能约 28%。", meas=True)

const("分层", "第1档生成期望 E", "E1", 2.2, "—", "关 1–2，生成池 2(90%)/4(10%)。该档技巧门槛：冲 256。")
const("分层", "第2档生成期望 E", "E2", 3.0, "—", "关 3–4，生成池 2(70%)/4(20%)/8(10%)。")
const("分层", "第3档生成期望 E", "E3", 4.8, "—", "关 5–6，生成池 4(80%)/8(20%)。该档技巧门槛：冲 512。")
const("分层", "第4档生成期望 E", "E4", 6.0, "—", "关 7–9，生成池 4(70%)/8(20%)/16(10%)。")
const("分层", "第5档生成期望 E", "E5", 9.6, "—", "关 10+，生成池 8(80%)/16(20%)。该档技巧门槛：冲 1024。")
const("分层", "棋盘跨关保留", "BOARD_CARRY", 1, "0 / 1", "★★ 1 = 保留。实测：弱分层 + 跨关保留下，750 步时熟练玩家 P(2048) = 48%，平庸 1.3%；不分层方案该值为 0%，主题会消失。", meas=True)

const("充能", "充能条上限", "CHARGE_MAX", 100, "80 – 150", "★ 附魔频率主旋钮。")
const("充能", "充能系数 k", "CHARGE_K", 2.9, "2.0 – 4.0", "充能' = k x log2(v/E) x LOG_CORR。k=2.9 → 约 1.5 次小附魔/波。v0.1 的 3.5 次/波是刻意的降频：小附魔频率不随技术变化，必须给超载让出输出份额。")
const("附魔", "基准附魔强度 EP", "EP_BASE", 180, "140 – 240", "★ v0.1 是 60。提高是因为附魔必须是主要伤害来源（支柱 P2）。")
const("附魔", "元素总伤害池系数", "ELEM_CAP", 2.5, "2.0 – 3.5", "★★ 单目标等效倍率。元素总伤害 = EP x 威力系数 x 此值，全场均分。v0.1 的火是「每个敌人各受 EPx1.0」= N x EP，13 敌时是雷的 5.2 倍 —— 已修正。")
const("附魔", "共鸣第2次加成", "RES2", 1.4, "1.2 – 1.6", "连续同元素第 2 次的威力倍率。")
const("附魔", "共鸣第3次+加成", "RES3", 1.8, "1.5 – 2.2", "第 3 次及以后。过低则轮盘编排无意义。")

const("经济", "星核系数", "K_STAR", 0.15, "0.08 – 0.22", "★ 星核 = CV' x 此值。v0.1 是 0.12，但那是配未归一化的 CV。目标：每波约 21 星核 ≈ 0.9 次战术指令。")
const("经济", "金币系数", "K_GOLD", 1.0, "0.6 – 1.5", "金币 = CV' x 此值 x 关卡倍率。")
const("经济", "碎片系数", "K_SHARD", 0.04, "0.02 – 0.08", "碎片 = CV' x 此值 x 关卡倍率。小数进池不丢。")
const("经济", "升华结晶倍率", "SUB_MULT", 3.0, "2.0 – 5.0", "升华时 CV' = (v/E) x 此值。")
const("经济", "自动升华阈值", "SUB_TH", 512, "256 / 512 / 1024", "≥ 此值的方块触发超载后自动升华。实测：random 策略 750 步后盘面占用 100%，升华是必需品而非优化项。")

const("关卡", "波次时长(秒)", "WAVE_T", 43, "30 – 60", "标准波平均时长（关 1 实际为 30/35/45/45/60，见 10 表）。")
const("关卡", "每关波数", "WAVE_N", 5, "4 – 6", "5 波一关，约 215 秒。若数据显示单局中位数 >15 分钟，压到 4 波。")
const("关卡", "敌人HP缩放底数", "HP_BASE", 1.55, "1.48 – 1.62", "★★ 难度主旋钮。v0.2 主输出移到附魔（卡牌驱动），需重新标定。当前配置输出：新手 4 关 / 中等 5 关 / 高手 6 关 / 满养成 9 关。")
const("关卡", "敌人HP超线性项", "HP_SLIN", 0.05, "0 – 0.10", "HP(n) = 基础 x 底数^(n-1) x (1 + 此项 x (n-1))。")
const("关卡", "每关敌人总数", "ENEMY_TOT", 60, "40 – 80", "5 波 x 约 12 只。用于漏怪模型估算。")
const("关卡", "平均撞星枢伤害", "AVG_DMG", 6, "4 – 8", "各类敌人加权平均。用于漏怪模型估算。")
const("关卡", "敌人速度缩放", "SPD_SCL", 0.04, "0.02 – 0.06", "速度增长必须慢，否则不可控。")
const("关卡", "敌人伤害缩放底数", "DMG_BASE", 1.25, "1.15 – 1.35", "伤害缩放应慢于 HP，避免星枢被秒。")
const("关卡", "关卡收益倍率增量", "RWD_INC", 0.3, "0.2 – 0.5", "关卡倍率 1.0, 1.3, 1.7, 2.2, 2.8 …（增量递增 0.1）")
const("关卡", "星枢HP", "NODE_HP", 100, "80 – 200", "★ 术语变更：v0.1 的「核心 HP」改称「星枢 HP」，与货币「晶核」区分。")
const("关卡", "关卡间星枢回复%", "NODE_REGEN", 0.20, "0 – 0.35", "鼓励推进。")

const("成长", "每关获取卡牌数", "CARD_N", 5, "3 – 6", "每关 5 波，每波清完给 1 张；与波数 1:1。")
const("成长", "单卡平均效果", "CARD_PWR", 0.12, "0.08 – 0.18", "★ 玩家战力成长主旋钮。")
const("成长", "卡牌效率衰减", "CARD_DECAY", 0.88, "0.80 – 0.95", "每关卡牌边际收益递减，制造自然撞墙点。")
const("成长", "植物基础DPS", "PLANT_DPS", 7, "5 – 9", "★ v0.1 是 10（焰心草 12）。下调是为了把主输出还给附魔。")
const("成长", "植物有效命中率", "HIT_RATE", 0.75, "0.6 – 0.9", "★ 新增。植物 DPS 并非全额命中（射程/目标切换/溢出），用于 10 表波次预算。")
const("成长", "初始编队位", "SLOT_BASE", 3, "2 – 4", "编队植物数量。")
const("成长", "每星DPS加成", "STAR_MULT", 0.20, "0.15 – 0.30", "每升 1 星 DPS 提升。5★ = 1.8x。")
const("成长", "每级DPS加成", "LV_MULT", 0.02, "0.015 – 0.03", "每级 DPS 提升。Lv50 = 1.98x。★ 压缩过大会碾压局内卡牌。")
const("成长", "伙伴位效率", "SLOT_EFF", 0.50, "0.4 – 0.7", "第 4/5/6 编队位上的伙伴按此效率计入战力。")

const("技术", "附魔占总输出比例", "ENCH_SHARE", 0.60, "0.5 – 0.7", "★★ v0.1 是 0.45，但按 v0.1 数值实算只有 0.21（植物抢了戏）。目标 0.60，见 10 表校验。")
const("技术", "超载占附魔输出比例", "OVER_SHARE", 0.60, "0.5 – 0.75", "★★ 新增。小附魔频率不随技术变化（差 6%），所以技巧只能挂在超载上。这是技术回报的唯一来源。")
const("技术", "新手超载次数比", "SK_NOVICE", 0.33, "0.2 – 0.5", "★★ 实测推导：random 策略相对高手的超载达成率。T3 档 P(512)：random 19.3% vs greedy 58.7%。", meas=True)
const("技术", "中等超载次数比", "SK_MID", 0.60, "0.4 – 0.8", "random 与 greedy 之间。注意：corner（固定方向流）实测比 random 还差 8%，是负技巧 —— 见风险 R9。", meas=True)
const("技术", "高手超载次数比", "SK_EXPERT", 1.00, "—", "基准。", meas=True)
const("技术", "小附魔技术倍率", "SK_CHARGE", 1.05, "1.0 – 1.10", "★★ 实测：小附魔频率 random 1.87 / greedy 1.99，仅差 6%。几乎不随技术变化。", meas=True)

const("风险", "失败保底比例", "B_FAIL", 0.40, "0.2 – 0.6", "★★ 冒险倾向主旋钮。0.4 → 继续阈值 p>43%。见 07 表。")
const("风险", "下一关收益/当前池", "R_RATIO", 0.80, "0.5 – 1.2", "由关卡收益倍率递增推导。")
const("风险", "满养成倍率", "MAX_UPG", 1.6, "1.3 – 2.0", "满养成相对零养成的战力倍率。用于漏怪模型。★ 超过 1.8 会碾压局内卡牌。")

const("养成", "养成树基础成本", "TREE_BASE", 40, "30 – 60", "星尘。")
const("养成", "养成树成本增长", "TREE_GROW", 1.30, "1.22 – 1.38", "★ 长线时长主旋钮。1.30 → 全树约 3–5 周。")
const("养成", "养成树最大等级", "TREE_MAX", 10, "8 – 15", "每分支等级上限。")

# ============================================================
# 02 合成与充能（已归一化）
# ============================================================
ws = wb.create_sheet("02_合成与充能")
ws["A1"] = "方块值 → 充能' / CV' / 各货币 / 附魔伤害（★ v0.2 已归一化：CV' = v / E）"
ws["A1"].font = TITLE_FONT
r = 3
r = sec(ws, r, "第一步：选择档位（改 B5 的 E 值切换，全表联动）")
head(ws, r, ["参数", "值", "说明"], [26, 12, 96])
r += 1
ws.cell(row=r, column=1, value="当前档位生成期望 E").border = BOX
c = ws.cell(row=r, column=2, value=2.2); c.fill = INPUT_FILL; c.border = BOX; c.font = Font(bold=True)
ws.cell(row=r, column=3, value="弱分层五档：2.2 / 3.0 / 4.8 / 6.0 / 9.6。改这个单元格即可切换档位查看换算。").border = BOX
E_REF = f"$B${r}"
r += 2

r = sec(ws, r, "弱分层五档参考（v0.2 M2）")
head(ws, r, ["档位", "关卡", "生成池", "E[生成]", "平均合成值", "该档技巧门槛", "合 256 的 CV'", "合 512 的 CV'"],
     [8, 12, 22, 11, 13, 20, 15, 15])
r += 1
for tier, lvl, pool, e, gate in [
    (1, "关 1–2", "2(90%)/4(10%)", "=CONSTE1", "冲 256（熟练 37%）"),
    (2, "关 3–4", "2(70%)/4(20%)/8(10%)", "=CONSTE2", "冲 256 稳定化"),
    (3, "关 5–6", "4(80%)/8(20%)", "=CONSTE3", "冲 512（熟练 58%）"),
    (4, "关 7–9", "4(70%)/8(20%)/16(10%)", "=CONSTE4", "冲 512 稳定化"),
    (5, "关 10+", "8(80%)/16(20%)", "=CONSTE5", "冲 1024（熟练 57%）"),
]:
    ef = {"=CONSTE1": CONST["E1"], "=CONSTE2": CONST["E2"], "=CONSTE3": CONST["E3"],
          "=CONSTE4": CONST["E4"], "=CONSTE5": CONST["E5"]}[e]
    ws.cell(row=r, column=1, value=tier).border = BOX
    ws.cell(row=r, column=2, value=lvl).border = BOX
    ws.cell(row=r, column=3, value=pool).border = BOX
    c = ws.cell(row=r, column=4, value=f"={ef}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=5, value=f"=D{r}*{CONST['AVG_MERGE_COEF']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    ws.cell(row=r, column=6, value=gate).border = BOX
    c = ws.cell(row=r, column=7, value=f"=256/D{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=8, value=f"=512/D{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    r += 1
r += 1
r = note(ws, r, "★ 归一化 CV' 的含义：它衡量「这次合成的难度」，不是数值大小。关 1 合出 256（占全局 78% 总价值）值 116 CV；关 10 合出 256 只是日常，只值 27 CV。同一份努力拿到同一份回报。")
r = note(ws, r, "★ 注意「合 256 的 CV'」随档位下降 —— 这是刻意的：高关卡的 256 不再稀缺，救济感由更高的星级（512/1024）接棒。")
r += 1

r = sec(ws, r, "换算总表（按上方 E 值计算）")
head(ws, r, ["方块值 v", "log₂v", "充能'", "是否超载", "品质", "威力系数",
             "附魔总伤害池", "CV'(合成)", "CV'(升华)", "星核", "晶核", "金币(关1)", "碎片(关1)"],
     [11, 8, 10, 10, 8, 10, 16, 11, 11, 9, 8, 11, 11])
r += 1

OVER = {256: (1, 2.0), 512: (2, 3.5), 1024: (3, 6.0), 2048: (4, 10.0), 4096: (5, 16.0)}
for i in range(2, 14):  # 4 ... 8192
    v = 2 ** i
    ws.cell(row=r, column=1, value=v).border = BOX
    ws.cell(row=r, column=2, value=i).border = BOX
    c = ws.cell(row=r, column=3, value=f"={CONST['CHARGE_K']}*LOG(A{r}/{E_REF},2)*{CONST['LOG_CORR']}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    star, mult = OVER.get(v, (0, ""))
    ws.cell(row=r, column=4, value=("超载" if v in OVER else "—")).border = BOX
    ws.cell(row=r, column=5, value=(f"{star}★" if v in OVER else "—")).border = BOX
    if v in OVER:
        ws.cell(row=r, column=6, value=mult).border = BOX
        c = ws.cell(row=r, column=7, value=f"={CONST['EP_BASE']}*F{r}*{CONST['ELEM_CAP']}")
        c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    else:
        ws.cell(row=r, column=6, value=1.0).border = BOX
        c = ws.cell(row=r, column=7, value=f"={CONST['EP_BASE']}*F{r}*{CONST['ELEM_CAP']}")
        c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=8, value=f"=A{r}/{E_REF}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=9, value=f"=IF(A{r}>={CONST['SUB_TH']},A{r}/{E_REF}*{CONST['SUB_MULT']},0)")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=10, value=f"=H{r}*{CONST['K_STAR']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    core = 8 if v >= 2048 else 3 if v >= 512 else 1 if v >= 128 else 0
    ws.cell(row=r, column=11, value=core).border = BOX
    c = ws.cell(row=r, column=12, value=f"=H{r}*{CONST['K_GOLD']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=13, value=f"=H{r}*{CONST['K_SHARD']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    r += 1

r += 1
r = note(ws, r, "★ 附魔总伤害池 = EP x 威力系数 x ELEM_CAP。这是「全场总伤害」，不是单体伤害 —— 火会均分给所有敌人，雷会集中在最前排。两者总量相等，区别纯化为分摊 vs 集中。")
r = note(ws, r, "★ 充能' 公式里的 LOG_CORR = 0.78 是实测的 Jensen 修正项。合成值分布右偏，E[log] < log(E[·])，直接用 log(均值) 会高估充能约 28%。")
r = note(ws, r, "设计校验：一次 2048 超载（4★）= 180 x 10 x 2.5 = 4,500 总伤害，是关 1 波 1 全波敌人总 HP（570）的 7.9 倍。这是玩家截图分享的 jackpot 时刻。")
r += 1

r = sec(ws, r, "共鸣倍率（连续同元素触发）")
head(ws, r, ["连续次数", "威力加成", "等效总伤害池", "说明"], [12, 12, 16, 78])
r += 1
for n, val, desc in [
    (1, 1.0, "首次触发，基准。"),
    (2, f"={CONST['RES2']}", "轮盘排 [X,X] 的收益。"),
    ("3+", f"={CONST['RES3']}", "附带「元素余韵」：持续类效果时长 +50%。轮盘排 [X,X,X] 是爆发构筑核心。"),
]:
    ws.cell(row=r, column=1, value=n).border = BOX
    c = ws.cell(row=r, column=2, value=val); c.fill = (CALC_FILL if isinstance(val, str) else INPUT_FILL); c.border = BOX
    c.number_format = "0.00"
    c = ws.cell(row=r, column=3, value=f"={CONST['EP_BASE']}*B{r}*{CONST['ELEM_CAP']}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    ws.cell(row=r, column=4, value=desc).border = BOX
    r += 1

# ============================================================
# 03 波次节奏模型（已归一化）
# ============================================================
ws = wb.create_sheet("03_波次节奏模型")
ws["A1"] = "波次节奏模型 —— 归一化后 CV'/波 应跨档位恒定（★ 这是 M1 的验收标准）"
ws["A1"].font = TITLE_FONT
head(ws, 3, ["关卡", "生成池", "E[生成]", "平均合成值", "每波步数", "合成次数/波",
             "CV'/波", "充能'/波", "小附魔/波", "星核/波", "金币/波", "碎片/波", "目标校验"],
     [7, 22, 10, 13, 11, 13, 11, 12, 12, 10, 11, 10, 30])

TIERS = [
    (1, "2(90%)/4(10%)", "E1"),
    (2, "2(70%)/4(20%)/8(10%)", "E2"),
    (3, "2(70%)/4(20%)/8(10%)", "E2"),
    (4, "4(80%)/8(20%)", "E3"),
    (5, "4(80%)/8(20%)", "E3"),
    (6, "4(70%)/8(20%)/16(10%)", "E4"),
    (7, "4(70%)/8(20%)/16(10%)", "E4"),
    (8, "4(70%)/8(20%)/16(10%)", "E4"),
    (9, "8(80%)/16(20%)", "E5"),
    (10, "8(80%)/16(20%)", "E5"),
    (12, "8(80%)/16(20%)", "E5"),
    (15, "8(80%)/16(20%)", "E5"),
]
r = 4
for stage, pool, ecode in TIERS:
    ws.cell(row=r, column=1, value=stage).border = BOX
    ws.cell(row=r, column=2, value=pool).border = BOX
    c = ws.cell(row=r, column=3, value=f"={CONST[ecode]}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=4, value=f"=C{r}*{CONST['AVG_MERGE_COEF']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=5, value=f"={CONST['WAVE_T']}/{CONST['STEP_REGEN']}+{CONST['STEP_GIFT']}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=6, value=f"=E{r}*{CONST['MERGE_RATE']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=7, value=f"=F{r}*{CONST['AVG_MERGE_COEF']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0"
    c = ws.cell(row=r, column=8, value=f"=F{r}*{CONST['CHARGE_K']}*LOG({CONST['AVG_MERGE_COEF']},2)*{CONST['LOG_CORR']}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0"
    c = ws.cell(row=r, column=9, value=f"=H{r}/{CONST['CHARGE_MAX']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=10, value=f"=G{r}*{CONST['K_STAR']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=11, value=f"=G{r}*{CONST['K_GOLD']}*(1+{CONST['RWD_INC']}*(A{r}-1))")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=12, value=f"=G{r}*{CONST['K_SHARD']}*(1+{CONST['RWD_INC']}*(A{r}-1))")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    ws.cell(row=r, column=13, value="CV' 应恒定 ≈ 151；小附魔 ≈ 1.5").border = BOX
    r += 1

r += 1
r = note(ws, r, "★★【M1 验收标准】CV'/波 与 小附魔/波 两列必须跨档位基本恒定。若出现随关卡单调上升，说明归一化失效，下游经济会重新膨胀。")
r = note(ws, r, "【对照 v0.1】未归一化时 CV/波 从 299 涨到 5,031（17 倍），星核从 25 涨到 604 —— 后期每波能放 40 次战术指令，战斗压力彻底失效。")
r = note(ws, r, "★【最根本的认知】2048 的合成产出是守恒量：每步生成 1 个方块，稳态必然消耗约 0.90 个（实测 0.89–0.94，跨策略差异 <6%）。所以总 CV ≈ 步数 x 生成值 x 常数，与技术水平无关。任何建立在「合成频率 / 合成总量」上的技巧表达都是无效的。", warn=True)
r = note(ws, r, "★【推论】技巧只能挂在「大数字达成率」上。小附魔频率实测 random 1.87 / greedy 1.99，仅差 6%；而 P(2048) 是 1.3% vs 48%。这就是为什么小附魔必须降频给超载让路。", warn=True)
r = note(ws, r, "【设计意图】星核/波 目标 20–30，即每波约 0.9 次战术指令。若超过 50，说明玩家能靠指令无限续命，战斗压力失效。")

# ============================================================
# 04 关卡与敌人
# ============================================================
ws = wb.create_sheet("04_关卡与敌人")
ws["A1"] = "关卡缩放 —— 敌人强度 vs 玩家战力成长（★ v0.2 换用大数字达成率驱动的技术模型）"
ws["A1"].font = TITLE_FONT
head(ws, 3, ["关卡 n", "敌人HP倍率", "敌人伤害倍率", "敌人速度倍率", "关卡收益倍率",
             "玩家战力倍率", "余量比", "状态", "第5波敌人总HP(估)", "玩家DPS(估)"],
     [8, 12, 13, 13, 12, 13, 10, 16, 16, 13])

r = 4
for n in range(1, 16):
    ws.cell(row=r, column=1, value=n).border = BOX
    c = ws.cell(row=r, column=2, value=f"={CONST['HP_BASE']}^(A{r}-1)*(1+{CONST['HP_SLIN']}*(A{r}-1))")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=3, value=f"={CONST['DMG_BASE']}^(A{r}-1)"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=4, value=f"=1+{CONST['SPD_SCL']}*(A{r}-1)"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    if n == 1:
        c = ws.cell(row=r, column=5, value=1.0); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
        c = ws.cell(row=r, column=6, value=1.0); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.000"
    else:
        c = ws.cell(row=r, column=5, value=f"=E{r-1}+{CONST['RWD_INC']}+0.1*(A{r}-2)")
        c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
        c = ws.cell(row=r, column=6, value=f"=F{r-1}*(1+{CONST['CARD_PWR']}*{CONST['CARD_DECAY']}^(A{r}-2))^{CONST['CARD_N']}")
        c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.000"
    c = ws.cell(row=r, column=7, value=f"=F{r}/B{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=8, value=f'=IF(G{r}>1.15,"轻松",IF(G{r}>0.85,"势均力敌",IF(G{r}>0.6,"吃力","撞墙")))')
    c.fill = CALC_FILL; c.border = BOX; c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r, column=9, value=f"=4330*B{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=10, value=f"=({CONST['PLANT_DPS']}*{CONST['SLOT_BASE']}*{CONST['HIT_RATE']})*F{r}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    r += 1

r += 1
r = note(ws, r, "【余量比判读】>1.15 轻松 / 0.85–1.15 势均力敌 / 0.6–0.85 吃力 / <0.6 撞墙。")
r = note(ws, r, "★ 余量比只是 DPS 上限，是乐观估计。玩家真正的死因不是「打不动」，而是「漏过去的敌人把星枢打爆」。必须用下面的漏怪模型判定实际死亡关卡。")
r += 1

r = sec(ws, r, "技术系数模型（★ v0.2 重写）")
head(ws, r, ["参数", "值", "说明"], [26, 12, 96])
r += 1
SKILL_ROW = {}
for label, code, desc in [
    ("新手 技术系数", "SK_NOVICE", "超载次数比 0.33（random 策略实测）"),
    ("中等 技术系数", "SK_MID", "超载次数比 0.60"),
    ("高手 技术系数", "SK_EXPERT", "超载次数比 1.00（基准）"),
]:
    ws.cell(row=r, column=1, value=label).border = BOX
    inner = (f"(1-{CONST['OVER_SHARE']})*{CONST['SK_CHARGE']}+{CONST['OVER_SHARE']}*{CONST[code]}")
    c = ws.cell(row=r, column=2, value=f"=(1-{CONST['ENCH_SHARE']})+{CONST['ENCH_SHARE']}*({inner})")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.000"
    ws.cell(row=r, column=3, value=desc).border = BOX
    SKILL_ROW[code] = f"$B${r}"
    r += 1
ws.cell(row=r, column=1, value="技术回报（高手 ÷ 新手）").border = BOX
c = ws.cell(row=r, column=2, value=f"={SKILL_ROW['SK_EXPERT']}/{SKILL_ROW['SK_NOVICE']}")
c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"; c.font = Font(bold=True, color="C00000")
ws.cell(row=r, column=3, value="★ v0.1 假设 1.65x，实测修正后约 1.31x。原因是植物（40%）与小附魔（24%）都不随技术变化，稀释了超载的贡献。").border = BOX
r += 2
r = note(ws, r, "★【v0.2 文档勘误】GDD v0.2 §1 M3 写的「总技术回报 1.62x」是错的 —— 它把「超载占附魔 60%」误当成「超载占总输出 60%」代入公式。正确值见上表，约 1.31x。", warn=True)
r = note(ws, r, "【技术模型结构】总输出 = 植物 40%（技术无关）+ 小附魔 24%（技术无关，倍率 1.05）+ 超载 36%（技术强相关）。要让技术回报继续上升，只能提高 OVER_SHARE 或 ENCH_SHARE，代价是植物存在感进一步下降。")
r += 1

r = sec(ws, r, "漏怪预算模型 —— 真实死亡关卡判定")
r = note(ws, r, "逻辑：当余量比 x 技术系数 x 养成系数 < 1 时，缺口比例的敌人会漏过防线；漏怪数 x 敌人伤害 > 星枢 HP 即阵亡。")
head(ws, r, ["关卡 n", "余量比(基准)", "可承受漏怪数", "新手", "中等", "高手", "满养成高手"], [8, 13, 14, 14, 14, 14, 15])
r += 1
for n in range(1, 13):
    src_row = n + 3
    ws.cell(row=r, column=1, value=n).border = BOX
    c = ws.cell(row=r, column=2, value=f"=G{src_row}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=3, value=f"={CONST['NODE_HP']}/({CONST['AVG_DMG']}*{CONST['DMG_BASE']}^(A{r}-1))")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    for idx, (sk, upg) in enumerate([("SK_NOVICE", 1), ("SK_MID", 1), ("SK_EXPERT", 1), ("SK_EXPERT", CONST["MAX_UPG"])], start=4):
        f = f'=IF({CONST["ENEMY_TOT"]}*MAX(0,1-B{r}*{SKILL_ROW[sk]}*{upg})<C{r},"存活","阵亡")'
        c = ws.cell(row=r, column=idx, value=f); c.fill = CALC_FILL; c.border = BOX
        c.alignment = Alignment(horizontal="center")
    r += 1
r += 1
r = note(ws, r, "【模型输出（HP_BASE=1.55，v0.2 技术模型）】新手 4 关 / 中等 5 关 / 高手 6 关 / 满养成高手 9 关。对比 v0.1 的 3/6/7/9 —— 高手端下移 1 关，因为 v0.2 高估的技术回报被修正掉了。")
r = note(ws, r, "【调参方向】若希望高手回到 7 关：把 HP_BASE 降到 1.50，或提高 OVER_SHARE 至 0.70（代价是植物存在感下降）。若新手 4 关仍挫败：前 2 关波次时长 +10s。")
r += 1

r = sec(ws, r, "敌人基准（第 1 关）★ v0.2 全量重算")
head(ws, r, ["类型", "HP", "v0.1 旧值", "速度(格/秒)", "撞星枢伤害", "护甲", "掉落材料", "设计目的"], [14, 9, 11, 12, 12, 9, 10, 54])
r += 1
for row in [
    ("小兵 Grunt", 95, 40, 0.35, 5, 0, 0, "基准单位。v0.1 的 40 在植物 DPS 腰斩后会被瞬秒。"),
    ("群聚 Swarm", 25, 12, 0.50, 2, 0, 0, "惩罚单体流，奖励火元素（AoE 均分）。"),
    ("迅捷 Swift", 85, 22, 0.75, 3, 0, 0, "★ v0.1 的 22 严重偏低，构不成时间压力。"),
    ("重甲 Armor", 190, 140, 0.22, 12, 0.30, 0, "惩罚穿透/破甲不足，奖励雷元素（集中伤害）。"),
    ("破坏者 Breaker", 200, 200, 0.30, 15, 0.10, 0, "攻击并摧毁植物格，惩罚固守不动的编队。"),
    ("精英 Elite", 450, 400, 0.28, 25, 0.15, "1–2", "携带元素护盾，检验玩家的轮盘编排。"),
    ("Boss", 1400, 1200, 0.20, 40, 0.20, "3–5", "双元素护盾，关卡高潮。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i == 8: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
r = note(ws, r, "这些数值来自 10_波次预算表的完整配平，不是拍的。改任意一个都要回 10 表重算余量。")
r += 1

r = sec(ws, r, "波次组成（第 1 关）★ 与 10 表联动")
head(ws, r, ["波", "组成", "时长(秒)", "敌人总HP", "玩家总输出", "余量比", "设计意图"], [6, 34, 11, 12, 12, 10, 54])
r += 1
for row in [
    (1, "6 小兵", 30, 570, 967, 1.70, "教学波。不可能失败。"),
    (2, "4 小兵 + 4 群聚 + 2 迅捷", 35, 612, 1121, 1.83, "引入群体压力与时间压力。"),
    (3, "2 重甲 + 6 小兵", 45, 950, 1428, 1.50, "引入护甲，制造第一次「打不动」的挫败。"),
    (4, "6 迅捷 + 2 重甲", 45, 890, 1428, 1.60, "时间压力为主，逼玩家加快合成。"),
    (5, "1 精英(火盾) + 8 小兵 + 4 群聚", 60, 1310, 1889, 1.44, "Boss 波。首次检验轮盘编排是否针对护盾。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i == 7: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
r = note(ws, r, "关 1 总 HP ≈ 4,332，总时长 215s，总步数 153。整关余量比约 1.6 —— 这是刻意的爽感期。")

# ============================================================
# 05 植物
# ============================================================
ws = wb.create_sheet("05_植物")
ws["A1"] = "植物数值与成长模型（★ v0.2：植物降级为辅助定位，主输出交给附魔）"
ws["A1"].font = TITLE_FONT
r = note(ws, 2, "成长公式：DPS = 基础 × (1 + 每星DPS加成×(星级−1)) × (1 + 每级DPS加成×(等级−1)) × 养成树枝叶加成")
r = note(ws, 3, "★ v0.2 变更：基础 DPS 整体下调约 42%（焰心草 12 → 7）。原因：按 v0.1 数值实算，关 1 每波植物输出是附魔的 2.7 倍，玩家不滑棋盘也能赢，直接违反支柱 P2。", warn=True)
head(ws, 4, ["植物", "元素", "攻击形态", "基础DPS", "v0.1 旧值", "射程(格)", "特性", "大招(4晶核/25s)"], [13, 8, 13, 10, 11, 10, 42, 44])
r = 5
for row in [
    ("焰心草", "火", "直射单体", 7, 12, 3, "命中附加灼烧：3 秒内每秒 3 伤害", "全场灼烧 8s，每秒 EP×0.15"),
    ("潮汐藤", "水", "环形 AoE", 4, 7, 1.5, "范围内敌人 −20% 移速（持续）", "全体击退 2 格 + 减速 50%（4s）"),
    ("荆棘木", "木", "穿透直线", 6, 11, "无限", "无视前排阻挡，贯穿本行", "贯穿全屏光束，EP×3.0"),
    ("辉光苔", "光", "辅助(无攻击)", 0, 0, "—", "每 10s：+1 步 或 +8 充能（玩家预设）", "立即 +4 步 +60 充能，全植物攻速 +50%（6s）"),
    ("雷鸣花", "雷", "溅射", 6, 10, 2.5, "命中溅射周围 50% 伤害", "15 次随机落雷，每次 EP×0.4"),
    ("霜晶莲", "冰", "直射单体", 5, 8, 3, "命中减速 30%（2s）", "全场冻结 4s，冻结期间受伤 +25%"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 7: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "星级 / 等级成长曲线（以荆棘木 基础DPS=6 为例）")
head(ws, r, ["等级\\星级", "1★", "2★", "3★", "4★", "5★"], [12, 11, 11, 11, 11, 11])
r += 1
for lv in [1, 10, 20, 30, 40, 50]:
    c = ws.cell(row=r, column=1, value=f"Lv.{lv}"); c.border = BOX
    for s in range(1, 6):
        f = f"=6*(1+{CONST['STAR_MULT']}*({s}-1))*(1+{CONST['LV_MULT']}*({lv}-1))"
        c = ws.cell(row=r, column=s + 1, value=f); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    r += 1
r += 1
r = note(ws, r, "【设计校验】Lv.50 / 5★ 的荆棘木 = 6 × 1.8 × 1.98 = 21.4 DPS，是初始值的 3.56 倍（倍率不变，只是基数下调）。配合编队位扩展至 6（伙伴位 50% 效率 → ×1.5），植物侧满养成合计约 5.3 倍。")
r = note(ws, r, "★★【平衡红线 · 最重要的一个】养成 vs 局内卡牌的战力占比，建议 60 : 40。监控方式：分别统计「满养成 + 零卡牌」与「零养成 + 满卡牌」两种极端配置下的可达关卡，二者之比应落在 1.3 – 1.8 之间。")
r += 1

r = sec(ws, r, "六元素效果表 ★ 火的 AoE 已修正（v0.2 M4）")
head(ws, r, ["元素", "定位", "总伤害池", "效果（0★ / EP=180）", "设计意图"], [8, 12, 12, 50, 50])
r += 1
for row in [
    ("火", "群体爆发", "≈EP×2.5", "★ 全场均分 EP×2.5 总伤害；已灼烧目标 +30%", "清群首选。v0.1 是「每个敌人各受 EP×1.0」= N×EP，13 敌时是雷的 5.2 倍 —— 已封顶修正。"),
    ("雷", "单体爆发", "≈EP×2.5", "连锁闪电弹跳 5 次，每次 EP×0.5，优先最前排", "打 Boss / 重甲。与火总伤害相等，区别纯化为分摊 vs 集中。"),
    ("冰", "控制+增伤", "—", "全场定身 3s，冻结目标受伤 +25%", "救场技，为爆发铺路。进攻向。"),
    ("水", "持续减速", "—", "移速 −40%，持续 5s，造成伤害 −20%", "拖时间，给棋盘争取步数。防守向。"),
    ("木", "己方增益", "—", "植物攻速 +35%，持续 8s；星枢回复 EP×0.2", "长线收益，配合高 DPS 编队才有价值。"),
    ("光", "资源调度", "—", "清除棋盘所有 ≤4 方块；+2 步；植物伤害 +25%（8s）", "唯一直接强化棋盘的元素，是解卡保底。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 4: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
r = note(ws, r, "★ 平衡校验：火与雷总伤害池都是 EP × 2.5 = 450（EP=180）。火分摊给全场（对 6 敌每人 75），雷集中在最前排（5 跳 × 90）。场面宽选火，打 Boss 选雷 —— 无绝对优劣。")
r += 1
r = sec(ws, r, "元素护盾（让轮盘编排有外部约束）")
head(ws, r, ["护盾类型", "效果", "破解方式", "设计目的"], [16, 40, 30, 46])
r += 1
for row in [
    ("元素护盾 · X", "受到非 X 元素伤害 −60%", "用 X 元素附魔 / X 属性植物攻击", "防止玩家把轮盘排成固定最优解。"),
    ("元素易伤 · X", "受到 X 元素伤害 +100%", "针对性编排轮盘", "奖励开局前的配置决策。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ============================================================
# 06 卡牌（新增战力点）
# ============================================================
ws = wb.create_sheet("06_卡牌")
ws["A1"] = "卡池清单 · 稀有度权重 · ★ 战力点等价（v0.2 新增配平工具）"
ws["A1"].font = TITLE_FONT
r = sec(ws, 2, "稀有度权重（按关卡）")
head(ws, 3, ["关卡", "普通", "稀有", "史诗", "传说", "说明"], [8, 10, 10, 10, 10, 60])
r = 4
for row in [
    (1, 0.70, 0.25, 0.05, 0.00, "教学期，几乎只有基础卡。"),
    (2, 0.60, 0.30, 0.09, 0.01, "首次见到传说，制造惊喜。"),
    (3, 0.50, 0.33, 0.15, 0.02, "构筑开始成型。"),
    (4, 0.42, 0.35, 0.18, 0.05, "构筑发力期。"),
    ("5+", 0.35, 0.35, 0.22, 0.08, "后期。同关第 4、5 波再 +5% 稀有度权重。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if 2 <= i <= 5: c.number_format = "0%"
        if i == 6: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "战力点基准（1 战力点 = 关 1 每波 1% 输出 = 11.5 伤害）")
head(ws, r, ["项目", "每波数值", "占比", "说明"], [26, 13, 10, 80])
r += 1
for label, val, share, desc in [
    ("关 1 每波总输出", 1148, "100%", "见 10_波次预算表。这是战力点的分母。"),
    ("其中 · 附魔", 675, "59%", "小附魔 1.5 次 × EP 180 × ELEM_CAP 2.5。"),
    ("其中 · 植物", 473, "41%", "3 株 × DPS 7 × 30s × HIT_RATE 0.75。"),
    ("1 战力点 =", 11.5, "1%", "所有卡牌效果都折算到这把尺子上。"),
]:
    ws.cell(row=r, column=1, value=label).border = BOX
    c = ws.cell(row=r, column=2, value=val); c.fill = MEAS_FILL; c.border = BOX; c.number_format = "#,##0.0"
    ws.cell(row=r, column=3, value=share).border = BOX
    ws.cell(row=r, column=4, value=desc).border = BOX
    r += 1
r += 1
r = note(ws, r, "★ 合格区间：普通 4–6 点 / 稀有 7–10 点 / 史诗 12–18 点 / 传说 20+ 点。新增卡牌一律先算战力点再定稀有度。")
r = note(ws, r, "★ 三条由此得出配平规则：① 步数/充能类卡是最强杠杆（同时放大 CV、充能、超载三条链路），不该出现在普通稀有度；② 经济卡战力点为 0，不该与战斗卡同池竞争，建议独立成池，每波第 3 次三选一固定从经济池抽；③ 战力点 < 2 的卡是废卡，等于卡池缩水。")
r += 1

r = sec(ws, r, "卡池（★ 战力点为实测折算，处置列为 v0.2 建议）")
head(ws, r, ["类别", "卡名", "稀有度", "效果", "战力点", "处置建议"], [10, 16, 9, 46, 9, 46])
r += 1
CARDS = [
    ("棋盘", "丰饶", "史诗", "每次移动额外生成 1 个方块", 58.0, "🔴 破坏性超模（正常卡的 10 倍）。重做或删除。"),
    ("棋盘", "充能加速", "普通", "每次合成充能 +25%", 14.7, "🔴 超模。升为史诗，或削到 +12%。"),
    ("棋盘", "疾风", "普通", "步数回复速度 +15%", 8.8, "🟡 偏强。升为稀有。"),
    ("棋盘", "精准", "稀有", "生成值为高档的概率 +25%", 6.5, "✅ 合格。"),
    ("棋盘", "聚能", "普通", "充能条上限 −10", 6.2, "✅ 合格（注意：充能降频后此卡价值同步下降）。"),
    ("棋盘", "净化术", "稀有", "立即清除所有 ≤8 方块并 +2 步", 4.0, "✅ 情境救急卡。"),
    ("棋盘", "稳健", "普通", "步数上限 +1", 2.0, "🔴 近废卡（步数回复才是瓶颈，不是容量）。改为「步数回复 +8%」。"),
    ("棋盘", "连锁核心", "传说", "一次移动中第 3 次及以后的合并，充能与 CV 翻倍", 0.5, "🔴 废卡且方向错误。实测 random 3连 10.3% > greedy 3.5%，此卡奖励乱滑。删除重做。"),
    ("战斗", "共生", "传说", "每个存活植物使全植物伤害 +8%", 9.8, "✅ 合格（3 株时）；满编队 6 株会超模，建议改为 +5%/株。"),
    ("战斗", "锋锐", "普通", "全植物伤害 +15%", 6.2, "✅ 合格，作为普通卡的战力基准。"),
    ("战斗", "暴击", "稀有", "暴击率 +10%，暴击伤害 150%", 5.0, "✅ 合格。"),
    ("战斗", "连射", "普通", "全植物攻速 +12%", 4.9, "✅ 合格。"),
    ("战斗", "破甲", "稀有", "无视 30% 护甲", 3.6, "🟡 情境卡。需提高其出现率，否则玩家觉得抽到废卡。"),
    ("战斗", "荆棘", "史诗", "敌人撞星枢时反伤 200", 3.0, "🟡 偏弱。后期压力大时才有用，建议按关卡动态加权。"),
    ("战斗", "远眺", "普通", "全植物射程 +1", 1.5, "🔴 废卡。依赖编队且收益不可见，重做。"),
    ("附魔", "共鸣强化", "稀有", "共鸣加成 1.4→1.6，1.8→2.2", 11.0, "✅ 合格（纯色队核心）。"),
    ("附魔", "超载", "史诗", "大附魔品质 +1 级（512 视为 1024）", 15.0, "✅ 合格。★ 这是提高技术回报最直接的卡，可考虑提出现率。"),
    ("附魔", "元素亲和 · X", "普通", "X 元素威力 +30%（六种各一张）", 2.9, "🟡 偏弱（只覆盖 1/6 的触发）。改为 +50%，或改「全元素 +15%」。"),
    ("附魔", "双生", "史诗", "附魔时同时触发轮盘下一位元素，威力各 60%", 12.0, "✅ 合格。"),
    ("附魔", "星爆", "传说", "2048 附魔威力 +100%，且必定暴击", 6.0, "🟡 看似传说实则情境（需先合到 2048）。建议叠加其他效果。"),
    ("经济", "贪婪", "普通", "金币 +25%", 0.0, "🟡 战力点 0，与战斗卡同池时永远不选。移入独立经济池。"),
    ("经济", "采集", "普通", "碎片产出 +50%", 0.0, "🟡 同上。移入独立经济池。"),
    ("经济", "结晶", "稀有", "所有 CV × 1.15", 0.3, "🟡 同上。移入独立经济池。"),
    ("经济", "升华", "稀有", "自动升华阈值降至 256", 8.0, "✅ 兑现流核心，有真实战力（腾盘面 + 提前变现）。可留在主池。"),
    ("经济", "囤积", "史诗", "大招晶核消耗 −1，星核产出 +20%", 4.5, "🟡 偏弱。提高星核部分至 +35%。"),
    ("战术", "时间膨胀", "稀有", "每波开始 +3 步", 8.8, "✅ 合格（等价于每波 +10% 步数）。"),
    ("战术", "战术家", "普通", "星核产出 +30%", 0.9, "🟡 偏弱。救急能力难以量化，建议改为「战术指令冷却 −20%」。"),
    ("战术", "急救", "稀有", "星枢 HP 首次低于 30% 时回复 40%", 5.0, "✅ 合格（一次性保险）。"),
    ("战术", "备用核心", "史诗", "星枢归零时以 30% HP 复活一次", 14.0, "✅ 合格。★ 实际价值远高于描述 —— 它等价于把 B_FAIL 临时抬高，会改变继续/收工的期望值。"),
    ("战术", "贪婪核心", "传说", "关闭自动升华", 18.0, "✅ 合格。高风险高回报的极限流开关，是升华机制的策略深度所在。"),
]
for row in CARDS:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i == 5:
            c.fill = MEAS_FILL
            c.number_format = "0.0"
            c.alignment = Alignment(horizontal="center")
            if v >= 20 or v < 2: c.font = Font(bold=True, color="C00000")
        if i in (4, 6): c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
r = note(ws, r, "卡池当前 30 张（含 6 张元素亲和变体则共 35 张）。V2 完整版目标 ≥ 60 张，否则 4–6 关后卡池枯竭，构筑深度崩塌。")
r = note(ws, r, "本次查出 4 张需删除/重做（丰饶、连锁核心、稳健、远眺）、2 张需升稀有度（充能加速、疾风）、1 张需降强度（共生）。处理后可用卡池约 26 张，需补充 34 张。")
r = note(ws, r, "平衡红线：同名卡重复时效果加法叠加而非乘法，防止指数爆炸。若埋点显示某卡选取率 > 70%，说明该卡超模或同类卡过少。")

# ============================================================
# 07 风险决策 EV
# ============================================================
ws = wb.create_sheet("07_风险决策EV")
ws["A1"] = "继续 / 收工 期望值计算器"
ws["A1"].font = TITLE_FONT
r = note(ws, 2, "公式：收工 EV = P ｜ 继续 EV = p×(P+R) + (1−p)×b×P ｜ 继续优于收工 ⟺ p > (1−b)P / ((1−b)P + R)")
r = 3
r = sec(ws, r, "主计算")
head(ws, r, ["参数", "值", "说明"], [30, 14, 74])
r += 1
ws.cell(row=r, column=1, value="当前累积收益池 P").border = BOX
c = ws.cell(row=r, column=2, value=1000); c.fill = INPUT_FILL; c.border = BOX
ws.cell(row=r, column=3, value="任意单位。EV 结论与 P 的绝对值无关，只与 R/P 比例有关。").border = BOX
P_ref = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="下一关预期收益 R").border = BOX
c = ws.cell(row=r, column=2, value=f"={P_ref}*{CONST['R_RATIO']}"); c.fill = CALC_FILL; c.border = BOX
ws.cell(row=r, column=3, value="R 与 P 成比例（R = P x R_RATIO）。R_RATIO 由关卡收益倍率递增推导。").border = BOX
R_ref = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="失败保底比例 b").border = BOX
c = ws.cell(row=r, column=2, value=f"={CONST['B_FAIL']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0%"
ws.cell(row=r, column=3, value="引用 01 核心常量 B_FAIL。").border = BOX
b_ref = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="玩家自估通关概率 p").border = BOX
c = ws.cell(row=r, column=2, value=0.60); c.fill = INPUT_FILL; c.border = BOX; c.number_format = "0%"
ws.cell(row=r, column=3, value="玩家主观判断。设计目标是让临界阈值足够低，从而鼓励冒险。").border = BOX
p_ref = f"$B${r}"; r += 1
r += 1

head(ws, r, ["指标", "值", "说明"], [30, 14, 74])
r += 1
ws.cell(row=r, column=1, value="收工 EV").border = BOX
c = ws.cell(row=r, column=2, value=f"={P_ref}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
ws.cell(row=r, column=3, value="确定性收益。").border = BOX
ev_stop = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="继续 EV").border = BOX
c = ws.cell(row=r, column=2, value=f"={p_ref}*({P_ref}+{R_ref})+(1-{p_ref})*{b_ref}*{P_ref}")
c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
ws.cell(row=r, column=3, value="p×(P+R) + (1−p)×b×P    （b = 保底返还比例）").border = BOX
ev_go = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="EV 差值（继续 − 收工）").border = BOX
c = ws.cell(row=r, column=2, value=f"={ev_go}-{ev_stop}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
c.font = Font(bold=True)
ws.cell(row=r, column=3, value="> 0 表示理性上应该选择继续。").border = BOX
r += 1

ws.cell(row=r, column=1, value="临界通关概率 p*").border = BOX
c = ws.cell(row=r, column=2, value=f"=(1-{b_ref})*{P_ref}/((1-{b_ref})*{P_ref}+{R_ref})")
c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0%"
c.font = Font(bold=True, color="C00000")
ws.cell(row=r, column=3, value="★ p > p* 时继续是正期望。这是整个风险系统的核心数字。").border = BOX
r += 2

r = sec(ws, r, "保底比例 b 的敏感性分析（★ 冒险倾向主旋钮）")
head(ws, r, ["保底 b", "临界概率 p*", "玩家行为预测", "适用产品目标"], [11, 14, 46, 46])
r += 1
SENS_START = r
for b in [0.0, 0.2, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0]:
    c = ws.cell(row=r, column=1, value=b); c.border = BOX; c.number_format = "0%"
    c = ws.cell(row=r, column=2, value=f"=(1-A{r})/((1-A{r})+{CONST['R_RATIO']})")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0%"
    r += 1
PRED = {
    0.0: ("极度保守，玩家会提前收工", "硬核产品，追求策略纯度"),
    0.2: ("偏保守，多数玩家 3–4 关收工", "中核产品，单局时长可控"),
    0.3: ("略保守", "—"),
    0.4: ("★ 平衡点，决策有张力", "★ 推荐：H5 休闲偏中核"),
    0.5: ("略激进", "—"),
    0.6: ("激进，玩家几乎不主动收工", "追求高重开率与广告曝光"),
    0.8: ("极度激进，收工按钮形同虚设", "纯广告变现产品"),
    1.0: ("无风险，继续永远最优", "决策消失，不推荐"),
}
rr = SENS_START
for b in [0.0, 0.2, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0]:
    pred, goal = PRED[b]
    c = ws.cell(row=rr, column=3, value=pred); c.border = BOX; c.alignment = Alignment(wrap_text=True, vertical="top")
    c = ws.cell(row=rr, column=4, value=goal); c.border = BOX; c.alignment = Alignment(wrap_text=True, vertical="top")
    if b == 0.4:
        for col in (1, 2, 3, 4):
            ws.cell(row=rr, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
    rr += 1
r = rr + 1
r = note(ws, r, "【关键洞察】保底比例 b 从 0.2 提到 0.6，临界概率从 50% 降到 33% —— 玩家会明显更爱冒险。这是调节「重开率」最廉价的旋钮，优先用它做 A/B 测。")
r = note(ws, r, "★【隐藏的美妙性质】R/P 比例随关卡自然变化：第 2 关约 1.55（p*≈28%，极度鼓励冒险）→ 第 6 关约 0.53（p*≈53%，必须谨慎）。游戏在早期自动怂恿冒进、后期自动转为审慎 —— 这是收益倍率递增结构带来的免费礼物，不要在调参时破坏它。")
r = note(ws, r, "【注意】卡牌「备用核心」等价于把 b 临时抬高，会显著改变持有该卡时的决策。它的实际价值远高于描述（战力点 14，但决策价值另计）。")
r = note(ws, r, "★【v0.2 商业化约束】广告复活只保留 50% 收益池，不能是 100%。否则玩家把「失败」当正常流程，43% 的临界概率会失效，继续/收工的张力消失。", warn=True)
r = note(ws, r, "【UI 强制要求】关卡结算界面必须同屏显示：当前池 / 下一关倍率与威胁等级 / 继续的预估收益 +R / 失败保底规则 / ★新增：当前棋盘缩略图 + 升华按钮（因为棋盘跨关保留，玩家必须看到自己携带的资产）。")

# ============================================================
# 08 养成与花园
# ============================================================
ws = wb.create_sheet("08_养成与花园")
ws["A1"] = "星尘养成树与花园产出模型"
ws["A1"].font = TITLE_FONT
r = sec(ws, 2, "养成树成本模型")
r = note(ws, r, "成本公式：Cost(n) = 基础成本 × 增长系数^(n−1)")
head(ws, r, ["等级", "单级成本(星尘)", "累计成本", "根系效果", "枝叶效果", "花蕾效果", "果实效果"], [8, 15, 13, 26, 26, 22, 32])
r += 1
for lv in range(1, 11):
    ws.cell(row=r, column=1, value=lv).border = BOX
    c = ws.cell(row=r, column=2, value=f"=ROUND({CONST['TREE_BASE']}*{CONST['TREE_GROW']}^(A{r}-1),0)"); c.fill = CALC_FILL; c.border = BOX
    c = ws.cell(row=r, column=3, value=(f"=B{r}" if lv == 1 else f"=C{r-1}+B{r}")); c.fill = CALC_FILL; c.border = BOX
    ws.cell(row=r, column=4, value="步数+0.2 / 回复+3% / 充能条−2%").border = BOX
    ws.cell(row=r, column=5, value="植物伤害+3% / 星枢HP+4%").border = BOX
    ws.cell(row=r, column=6, value="全元素威力+4%").border = BOX
    ws.cell(row=r, column=7, value="金币+4% / 碎片+6% / 花园产出+5%").border = BOX
    r += 1
ws.cell(row=r, column=1, value="满级合计").font = Font(bold=True)
c = ws.cell(row=r, column=3, value=f"=SUM(B{r-10}:B{r-1})*4"); c.fill = CALC_FILL; c.border = BOX; c.font = Font(bold=True)
ws.cell(row=r, column=4, value="4 分支 × 单分支累计").font = Font(bold=True)
r += 2
r = note(ws, r, "【当前配置】基础 40 / 增长 1.30 → 单分支满级约 1,705 星尘，全树约 6,820 星尘。早期花园产出 200–500/天，后期 2,000+/天 → 约 3–5 周长线目标。")
r = note(ws, r, "【调参方向】H5 休闲产品建议长线不超过 4 周。若超长，把增长系数降到 1.25（全树约 5,300）；若希望更长线支撑赛季，升到 1.35。")
r += 1

r = sec(ws, r, "花园产出模型")
head(ws, r, ["稀有度", "星尘/小时(基)", "30min", "2h", "6h", "6 盆全传说/6h"], [12, 14, 10, 10, 10, 18])
r += 1
for name, base in [("普通", 8), ("稀有", 15), ("史诗", 28), ("传说", 50)]:
    ws.cell(row=r, column=1, value=name).border = BOX
    c = ws.cell(row=r, column=2, value=base); c.fill = INPUT_FILL; c.border = BOX
    for col, hrs in [(3, 0.5), (4, 2), (5, 6)]:
        c = ws.cell(row=r, column=col, value=f"=B{r}*{hrs}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0"
    c = ws.cell(row=r, column=6, value=f"=B{r}*6*6"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    r += 1
r += 1
r = note(ws, r, "产出速率 = 稀有度基础 × (1 + 星级×0.1) × (1 + 养成树果实分支等级×0.05)")
r = note(ws, r, "离线计时上限 12 小时（防纯挂机，同时保证次日回归有收益 → 这是次留的核心钩子）。")
r = note(ws, r, "花盆：初始 2，最多 6。金币解锁成本 5,000 / 20,000 / 60,000 / 150,000。")

# ============================================================
# 10 波次预算表（新增）
# ============================================================
ws = wb.create_sheet("10_波次预算表")
ws["A1"] = "关 1 波次预算表 —— 玩家输出能力 vs 敌人 HP（★ v0.2 新增，是 04 表敌人基准数值的来源）"
ws["A1"].font = TITLE_FONT
r = note(ws, 2, "配平逻辑：先算玩家在该波能打出多少输出，再除以目标余量比，得到敌人总 HP 预算。所有敌人基准 HP 由此反推。")
r = note(ws, 3, "附魔输出 = 合成次数 × k × log₂(5.47) × 0.78 ÷ 100 × (EP × ELEM_CAP)　｜　植物输出 = 3 株 × DPS 7 × 时长 × HIT_RATE 0.75", warn=True)
head(ws, 4, ["波", "时长(秒)", "步数", "合成次数", "小附魔次数", "附魔输出", "植物输出",
             "玩家总输出", "目标余量比", "敌人总HP(算)", "实际配置HP", "偏差", "组成"],
     [5, 10, 9, 11, 12, 12, 12, 13, 12, 14, 13, 9, 40])
r = 5
for w, t, comp, act, margin in [
    (1, 30, "6 小兵", 570, 1.70),
    (2, 35, "4 小兵 + 4 群聚 + 2 迅捷", 612, 1.70),
    (3, 45, "2 重甲 + 6 小兵", 950, 1.50),
    (4, 45, "6 迅捷 + 2 重甲", 890, 1.60),
    (5, 60, "1 精英(火盾) + 8 小兵 + 4 群聚", 1310, 1.44),
]:
    ws.cell(row=r, column=1, value=w).border = BOX
    c = ws.cell(row=r, column=2, value=t); c.fill = INPUT_FILL; c.border = BOX
    c = ws.cell(row=r, column=3, value=f"=B{r}/{CONST['STEP_REGEN']}+{CONST['STEP_GIFT']}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=4, value=f"=C{r}*{CONST['MERGE_RATE']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=5, value=f"=D{r}*{CONST['CHARGE_K']}*LOG({CONST['AVG_MERGE_COEF']},2)*{CONST['LOG_CORR']}/{CONST['CHARGE_MAX']}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=6, value=f"=E{r}*{CONST['EP_BASE']}*{CONST['ELEM_CAP']}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=7, value=f"={CONST['PLANT_DPS']}*{CONST['SLOT_BASE']}*B{r}*{CONST['HIT_RATE']}")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=8, value=f"=F{r}+G{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=9, value=margin); c.fill = INPUT_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=10, value=f"=H{r}/I{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=11, value=act); c.fill = INPUT_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=12, value=f"=K{r}/J{r}-1"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0%"
    c.font = Font(bold=True, color="C00000")
    ws.cell(row=r, column=13, value=comp).border = BOX
    r += 1
ws.cell(row=r, column=1, value="合计").font = Font(bold=True)
c = ws.cell(row=r, column=8, value=f"=SUM(H5:H{r-1})"); c.fill = CALC_FILL; c.border = BOX; c.font = Font(bold=True); c.number_format = "#,##0"
c = ws.cell(row=r, column=10, value=f"=SUM(J5:J{r-1})"); c.fill = CALC_FILL; c.border = BOX; c.font = Font(bold=True); c.number_format = "#,##0"
c = ws.cell(row=r, column=11, value=f"=SUM(K5:K{r-1})"); c.fill = CALC_FILL; c.border = BOX; c.font = Font(bold=True); c.number_format = "#,##0"
c = ws.cell(row=r, column=12, value=f"=K{r}/J{r}-1"); c.fill = CALC_FILL; c.border = BOX; c.font = Font(bold=True); c.number_format = "0%"
r += 2

r = sec(ws, r, "输出占比校验（★ M3 的验收点）")
head(ws, r, ["项目", "关 1 全关合计", "占比", "目标", "判定"], [26, 16, 11, 12, 46])
r += 1
for label, val, target, judge in [
    ("附魔输出", "=SUM(F5:F9)", "60%", "低于 50% 说明支柱 P2 失守，玩家不滑棋盘也能赢"),
    ("植物输出", "=SUM(G5:G9)", "40%", "高于 50% 说明植物抢戏，需再降 PLANT_DPS"),
]:
    ws.cell(row=r, column=1, value=label).border = BOX
    c = ws.cell(row=r, column=2, value=val); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=3, value=f"=B{r}/($B${r}+$B${r+1})"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0%"
    ws.cell(row=r, column=4, value=target).border = BOX
    ws.cell(row=r, column=5, value=judge).border = BOX; ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
r = note(ws, r, "★ 这是 M3 的验收点。v0.1 的实算是 21% : 79%（植物碾压），v0.2 目标 60% : 40%。若实测偏离超过 ±5%，改 EP_BASE 或 PLANT_DPS 而不是改敌人 HP。")
r = note(ws, r, "【偏差列用法】|偏差| > 15% 的行会标红，说明「实际配置 HP」与「模型预算 HP」脱节，需要二选一调整。")
r = note(ws, r, "【关 1 刻意宽松】整关余量比约 1.6，前 3 关是「我很强」的爽感期。这个爽感期是刻意的 —— 它保证每局前 12 分钟都是正反馈，失败发生在玩家已经投入之后。")

# ============================================================
# 09 验收清单
# ============================================================
ws = wb.create_sheet("09_验收清单")
ws["A1"] = "平衡目标与 Broken 判定标准 —— playtest 前必读（v0.2 更新）"
ws["A1"].font = TITLE_FONT
r = note(ws, 2, "原则：先定义「什么算坏」，才能在 playtest 中认出它。所有阈值在此表预先约定，不做事后解释。")
r = 4
r = sec(ws, r, "健康指标（埋点目标）★ 灰色行为 v0.2 修订/新增项")
head(ws, r, ["类别", "指标", "目标区间", "低于下限的处置", "高于上限的处置"], [10, 32, 18, 42, 42])
r += 1
for row in [
    ("节奏", "每波步数消耗", "25 – 32 步", "步数回复过快或玩家不积极 → 降低回复速率", "步数不够用 → 提高回复速率"),
    ("节奏", "每波合成次数", "25 – 30 次", "棋盘太挤 → 提升生成分层", "棋盘过于宽松 → 提升难度"),
    ("节奏", "每波小附魔触发", "1.2 – 1.8 次", "★ v0.2 刻意降频（原 3–4）。低于 1.2 → 降 CHARGE_MAX", "高于 1.8 → 升 CHARGE_MAX。注意：频率过高会挤压超载的输出份额， destroys 技术回报"),
    ("节奏", "每波战术指令使用", "0.8 – 1.5 次", "星核产出不足 → 提高 K_STAR", "指令续命 → 降低 K_STAR（>2 必须处理）"),
    ("难度", "平均到达关卡", "4 – 6 关", "难度过高 → 降低 HP_BASE 至 1.50", "难度过低 → 升高 HP_BASE 至 1.60"),
    ("难度", "第 1 关失败率", "< 5%", "—（达标即可）", "教学波设计失败，需重做波次组成"),
    ("难度", "无养成新手到达", "4 关", "新手挫败 → 前 2 关波次时长 +10s", "元游戏无意义 → 前期难度上调"),
    ("难度", "满养成高手到达", "9 – 11 关", "长线目标不足 → 提高养成上限", "养成碾压 → 压缩 STAR_MULT / LV_MULT / SLOT_EFF"),
    ("难度", "收工关卡众数", "第 4 – 6 关", "玩家不敢冒险 → 提高保底 b", "玩家不敢收工 → 降低保底 b"),
    ("留存", "单局时长", "12 – 20 分钟", "内容不足 → 增加波数", "★ 疲劳 → 把 WAVE_N 从 5 压到 4（H5 休闲基准是 3–5 分钟）"),
    ("健康", "棋盘卡死触发率", "< 3%", "—（达标即可）", "棋盘太小或生成算法有问题，需修伪随机"),
    ("★新版", "输出占比（附魔 : 植物）", "60 : 40（±5）", "植物 > 50% → 支柱 P2 失守，再降 PLANT_DPS", "植物 < 30% → 植物沦为摆设，回调 PLANT_DPS"),
    ("★新版", "CV'/波 跨档位波动", "< 10%", "—（达标即可）", "归一化失效，检查 01 表 E1–E5 是否与 03 表一致"),
    ("★新版", "超载触发次数 / 关", "关1-2 ≥0.3 / 关5-6 ≥1.5 / 关10+ ≥3", "过低 → 弱分层增益不足，技巧无回报", "—（越高越好，说明玩家在成长）"),
    ("★新版", "P(2048) 长局达成率", "熟练 40–55% / 新手 <5%", "过低 → 2048 主题形同虚设，检查 BOARD_CARRY", "过高 → 主题廉价化，降低分层增益"),
    ("★新版", "火 / 雷 元素选取率比", "0.8 – 1.3", "火 < 0.8 → AoE 封顶过狠，回调 ELEM_CAP", "火 > 1.5 → ELEM_CAP 未生效，检查实现"),
    ("★新版", "跨关后开局盘面占用", "关3 <55% / 关5 <70%", "—（达标即可）", "过高 → 下调 SUB_TH，或启用「关卡开局半价升华」备用方案"),
]:
    isnew = row[0] == "★新版"
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if isnew: c.fill = PatternFill("solid", fgColor="DDEBF7")
        if i >= 3: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "Broken 判定（出现任一症状即为设计失败，需回头改系统而非改数值）")
head(ws, r, ["症状", "判定阈值", "根因假设", "处置"], [30, 22, 42, 44])
r += 1
for row in [
    ("战斗毫无压力", "星枢 HP 全损率 < 10%", "敌人推进速度或数量不足", "提高 SPD_SCL，或在 10 表降低目标余量比"),
    ("棋盘被忽略", "步数常年满仓（余量 > 60%）", "合成收益不足以驱动玩家操作", "提高 CV 转化系数，或降低 STEP_REGEN"),
    ("附魔刷屏", "每波小附魔 > 2.5 次", "★ v0.2 阈值下调。充能系数过高", "降低 CHARGE_K 或升高 CHARGE_MAX"),
    ("卡牌无选择", "某张卡选取率 > 70%", "该卡超模或同类卡过少", "查 06 表战力点，超 20 点的一律削"),
    ("经济通胀", "金币日增量 / 日消耗 > 3", "金币汇不足", "增设金币汇，或降低 K_GOLD"),
    ("元素单一最优解", "某元素使用率 > 50%", "该元素超模，或克制设计未生效", "调整该元素，或增加对应护盾精英出现率"),
    ("后期棋盘饱和", "第 8 关后棋盘空格 < 3", "升华阈值过高", "降低 SUB_TH 至 256，或提高「贪婪核心」出现率"),
    ("养成碾压局内", "满养成 vs 零养成可达关卡比 > 1.8", "植物星级/等级/编队位成长过大", "压缩 STAR_MULT / LV_MULT / SLOT_EFF"),
    ("局内卡牌无意义", "满卡牌 vs 零卡牌可达关卡比 < 1.3", "卡牌效果或数量不足", "提高 CARD_PWR，或按 06 表建议重做废卡"),
    ("★技巧无回报", "高手 vs 新手可达关卡比 < 1.15", "超载输出占比不足（OVER_SHARE 过低）", "提高 OVER_SHARE 至 0.70，代价是植物存在感下降"),
    ("★固定方向流泛滥", "玩家连续同方向移动占比 > 40%", "★ 实测：corner 策略在 5×5 比 random 还差 8%，是负技巧", "新手引导需干预（GDD v0.2 §6-A）：连续 4 次同方向时触发预警提示"),
]:
    isnew = row[0].startswith("★")
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if isnew: c.fill = PatternFill("solid", fgColor="DDEBF7")
        if i >= 3: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "平衡流程（不可跳过）")
head(ws, r, ["阶段", "活动", "产出", "准入条件"], [18, 48, 34, 34])
r += 1
for row in [
    ("1. 纸面模拟", "★ 已完成。4 组蒙特卡洛校准（MERGE_RATE / 归一化 / 连锁 / 分层方案）", "校准后的 MERGE_RATE 0.90、AVG_MERGE_COEF 5.47、LOG_CORR 0.78", "✅ 完成"),
    ("2. 数值原型", "只做棋盘 + 充能 + 附魔伤害的裸 Demo（无战斗）", "验证乐趣假设：精准合成触发超载清屏的瞬间是否成立", "纸面模拟完成"),
    ("3. 可玩原型", "加入战斗与波次", "验证压力曲线与注意力分配", "乐趣假设成立"),
    ("4. 小规模测试", "20–50 人试玩，只观察 Feel 问题，不调数值", "Feel 问题清单", "核心循环可玩"),
    ("5. 数值调优", "基于埋点做回归分析", "v0.3 数值", "埋点体系上线"),
    ("6. 软启动 A/B", "测三个主旋钮：STEP_REGEN / B_FAIL / HP_BASE", "上线数值配置", "数值基本稳定"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 2: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 2
r = sec(ws, r, "给开发团队的一句话")
c = ws.cell(row=r, column=1, value="V1 垂直切片是生死线。如果 V1（1 主植物 / 3 波 / 2 元素 / 10 张卡）玩起来不爽，加再多内容和养成也救不回来 —— 请在 V1 投入充分的试玩时间，而不是赶着堆内容。")
c.font = Font(bold=True, size=11, color="C00000")
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.row_dimensions[r].height = 40

# 冻结首行
for s in wb.worksheets:
    s.freeze_panes = "A4" if s.title != "00_说明" else "A3"

# 调整 sheet 顺序
order = ["00_说明", "01_核心常量", "02_合成与充能", "03_波次节奏模型", "04_关卡与敌人",
         "05_植物", "06_卡牌", "07_风险决策EV", "08_养成与花园", "09_验收清单", "10_波次预算表"]
wb._sheets = [wb[n] for n in order]

out = r"C:\Users\creat\WorkBuddy AI\2026-08-31-15-18-31\Balance_星序防线_v0.2.xlsx"
wb.save(out)
print("SAVED:", out)
