# -*- coding: utf-8 -*-
"""生成《星序：元素防线》调参总表 v0.1 —— 所有数值联动，改核心常量即可全局重算。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ---------- 样式 ----------
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")   # 黄 = 可调输入
CALC_FILL  = PatternFill("solid", fgColor="EAF3DE")   # 绿 = 公式计算
HEAD_FILL  = PatternFill("solid", fgColor="4472C4")
HEAD_FONT  = Font(color="FFFFFF", bold=True, size=10)
SEC_FONT   = Font(bold=True, size=11, color="1F3864")
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
NOTE_FONT  = Font(size=9, color="808080", italic=True)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def head(ws, row, headers, widths=None):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEAD_FILL, HEAD_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BOX
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30

def put(ws, row, values, fill=None, numfmt=None, wrap=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = BOX
        if fill: c.fill = fill
        if numfmt and i > 1: c.number_format = numfmt
        if wrap: c.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1

def sec(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = SEC_FONT
    return row + 1

def note(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    return row + 1

# ============================================================
# 00 说明
# ============================================================
ws = wb.active
ws.title = "00_说明"
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 110
ws["A1"] = "星序：元素防线 — 调参总表 v0.1"
ws["A1"].font = TITLE_FONT
r = 3
r = sec(ws, r, "使用方式")
for a, b in [
    ("黄色单元格", "可调输入参数。改动后全表公式自动重算。"),
    ("绿色单元格", "公式计算结果。请勿手动覆盖，否则断链。"),
    ("灰色小字", "设计理由 / 调参方向说明。"),
    ("无底色", "静态参考数据（文案表、卡池清单等）。"),
]:
    ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=10)
    c = ws.cell(row=r, column=2, value=b); c.font = Font(size=10)
    r += 1
r += 1
r = sec(ws, r, "工作表导航")
for a, b in [
    ("01_核心常量", "全局旋钮。所有其他表的输入源头，优先在这里调。"),
    ("02_合成与充能", "方块值 → 充能 / CV / 各货币 / 附魔伤害的换算表。"),
    ("03_波次节奏模型", "估算每波的步数、合成次数、ΣCV、附魔次数、星核产出。"),
    ("04_关卡与敌人", "1–15 关的敌人强度缩放 vs 玩家战力成长，含「撞墙点」判定。"),
    ("05_植物", "植物基础数值与星级/等级成长公式。"),
    ("06_卡牌", "卡池清单与稀有度权重。"),
    ("07_风险决策EV", "继续 / 收工的期望值计算器与保底比例敏感性分析。"),
    ("08_养成与花园", "星尘养成树成本模型与花园产出模型。"),
    ("09_验收清单", "平衡目标与「broken」判定标准 —— playtest 前先读这张。"),
]:
    ws.cell(row=r, column=1, value=a).font = Font(bold=True, size=10)
    ws.cell(row=r, column=2, value=b).font = Font(size=10)
    r += 1
r += 1
r = note(ws, r, "所有数值均为 v0.1 假设，未经 playtest 验证。纸面模拟（10,000 局 Monte Carlo）是调参的第一步，不是可选项。")
r = note(ws, r, "配套文档：GDD_星序防线_v0.1.md")

# ============================================================
# 01 核心常量
# ============================================================
ws = wb.create_sheet("01_核心常量")
ws["A1"] = "核心常量 —— 全局旋钮（黄底可调，改这里会驱动全表）"
ws["A1"].font = TITLE_FONT
head(ws, 3, ["分组", "参数名", "代号", "当前值", "建议范围", "调参说明"], [14, 26, 14, 12, 16, 62])

CONST = {}
r = 4
def const(group, name, code, val, rng, why):
    global r
    ws.cell(row=r, column=1, value=group).border = BOX
    ws.cell(row=r, column=2, value=name).border = BOX
    ws.cell(row=r, column=3, value=code).border = BOX
    c = ws.cell(row=r, column=4, value=val); c.fill = INPUT_FILL; c.border = BOX; c.font = Font(bold=True)
    ws.cell(row=r, column=5, value=rng).border = BOX
    c = ws.cell(row=r, column=6, value=why); c.border = BOX; c.alignment = Alignment(wrap_text=True, vertical="top")
    CONST[code] = f"'01_核心常量'!$D${r}"
    r += 1
    return f"'01_核心常量'!$D${r-1}"

const("棋盘", "棋盘尺寸", "BOARD_N", 5, "4 / 5 / 6", "5×5。4×4 在实时压力下死局率过高；6×6 太宽松失去张力。")
const("棋盘", "每次移动生成方块数", "SPAWN_N", 1, "1 – 2", "保持 1。卡牌「丰饶」可临时 +1。")
const("棋盘", "初始方块数", "INIT_N", 2, "2 – 4", "留出前期铺垫空间。")
const("棋盘", "步数上限", "STEP_MAX", 5, "3 – 10", "攒满 5 步可打一套爆发连招，这是「攒劲儿」的爽点来源。")
const("棋盘", "步数回复间隔(秒)", "STEP_REGEN", 1.5, "1.0 – 2.5", "★ 手感主旋钮。直接决定每波步数预算与合成次数。")
const("棋盘", "波次开局赠送步数", "STEP_GIFT", 2, "0 – 4", "缓解波次切换空窗。")
const("棋盘", "每步平均合成次数", "MERGE_RATE", 1.2, "0.8 – 1.8", "★ 模型关键假设，必须用 Monte Carlo 校准。新手 0.5，高手 1.8+。")

const("充能", "充能条上限", "CHARGE_MAX", 100, "80 – 150", "★ 附魔频率主旋钮。目标：每波 3–4 次小附魔。")
const("充能", "充能系数 k", "CHARGE_K", 4, "2 – 8", "充能 = k × (log₂v − 1)。")
const("附魔", "基准附魔强度 EP", "EP_BASE", 60, "40 – 100", "对标第 1 关小兵 HP 40。所有元素效果的乘数基准。")
const("附魔", "共鸣第2次加成", "RES2", 1.4, "1.2 – 1.6", "连续同元素第 2 次的威力倍率。")
const("附魔", "共鸣第3次+加成", "RES3", 1.8, "1.5 – 2.2", "第 3 次及以后。过低则轮盘编排无意义。")

const("经济", "星核系数", "K_STAR", 0.12, "0.06 – 0.20", "★ 星核 = CV × 此值。目标：每波约 25 星核 = 1 次战术指令。")
const("经济", "金币系数", "K_GOLD", 1.0, "0.6 – 1.5", "金币 = CV × 此值 × 关卡倍率。")
const("经济", "碎片系数", "K_SHARD", 0.04, "0.02 – 0.08", "碎片 = CV × 此值 × 关卡倍率。小数进池不丢。")
const("经济", "升华结晶倍率", "SUB_MULT", 3.0, "2.0 – 5.0", "升华时 CV = v × 此值。过低则无人升华，过高则无人冲高星。")
const("经济", "自动升华阈值", "SUB_TH", 512, "256 / 512 / 1024", "≥ 此值的方块触发超载后自动升华。")

const("关卡", "波次时长(秒)", "WAVE_T", 40, "30 – 60", "标准波。Boss 波约为 1.5 倍。")
const("关卡", "每关波数", "WAVE_N", 5, "3 – 6", "5 波一关，约 4 分钟。")
const("关卡", "敌人HP缩放底数", "HP_BASE", 1.55, "1.40 – 1.70", "★★ 难度主旋钮。1.55 经漏怪模型校准：中等玩家约第 6 关阵亡。")
const("关卡", "敌人HP超线性项", "HP_SLIN", 0.05, "0 – 0.10", "HP(n) = 基础 × 底数^(n−1) × (1 + 此项 × (n−1))。")
const("关卡", "每关敌人总数", "ENEMY_TOT", 60, "40 – 80", "5 波 × 约 12 只。用于漏怪模型估算。")
const("关卡", "平均撞核心伤害", "AVG_DMG", 6, "4 – 8", "各类敌人加权平均。用于漏怪模型估算。")
const("关卡", "敌人速度缩放", "SPD_SCL", 0.04, "0.02 – 0.06", "速度增长必须慢，否则不可控。")
const("关卡", "敌人伤害缩放底数", "DMG_BASE", 1.25, "1.15 – 1.35", "伤害缩放应慢于 HP，避免核心被秒。")
const("关卡", "关卡收益倍率增量", "RWD_INC", 0.3, "0.2 – 0.5", "关卡倍率 1.0, 1.3, 1.7, 2.2, 2.8 …（增量递增 0.1）")
const("关卡", "核心HP", "CORE_HP", 100, "80 – 200", "防线血量。")
const("关卡", "关卡间核心回复%", "CORE_REGEN", 0.20, "0 – 0.35", "鼓励推进。")

const("成长", "每关获取卡牌数", "CARD_N", 5, "3 – 6", "= 每关波数。")
const("成长", "单卡平均效果", "CARD_PWR", 0.12, "0.08 – 0.18", "★ 玩家战力成长主旋钮。")
const("成长", "卡牌效率衰减", "CARD_DECAY", 0.88, "0.80 – 0.95", "每关卡牌边际收益递减，制造自然撞墙点。")
const("成长", "植物基础DPS", "PLANT_DPS", 10, "8 – 14", "1★1 级单体基准。")
const("成长", "初始编队位", "SLOT_BASE", 3, "2 – 4", "编队植物数量。")
const("成长", "每星DPS加成", "STAR_MULT", 0.20, "0.15 – 0.30", "每升 1 星 DPS 提升。5★ = 1.8×。")
const("成长", "每级DPS加成", "LV_MULT", 0.02, "0.015 – 0.03", "每级 DPS 提升。Lv50 = 1.98×。★ 压缩过大会碾压局内卡牌。")
const("成长", "伙伴位效率", "SLOT_EFF", 0.50, "0.4 – 0.7", "第 4/5/6 编队位上的伙伴植物按此效率计入战力（伙伴偏辅助向）。")
const("成长", "附魔占输出比例", "ENCH_SHARE", 0.45, "0.3 – 0.6", "★ 技术水平对总输出的影响权重。见 04 表漏怪模型。")
const("成长", "新手合成率", "MR_NOVICE", 0.5, "—", "新手每步平均合成次数（对比 MERGE_RATE 1.2）。")
const("成长", "高手合成率", "MR_EXPERT", 1.8, "—", "高手每步平均合成次数。")

const("风险", "失败保底比例", "B_FAIL", 0.40, "0.2 – 0.6", "★★ 冒险倾向主旋钮。0.4 → 继续阈值 p>43%。见 07 表。")
const("风险", "下一关收益/当前池", "R_RATIO", 0.80, "0.5 – 1.2", "由关卡收益倍率递增推导。")

const("养成", "养成树基础成本", "TREE_BASE", 40, "30 – 60", "星尘。")
const("养成", "养成树成本增长", "TREE_GROW", 1.30, "1.22 – 1.38", "★ 长线时长主旋钮。1.30 → 全树约 3–5 周。")
const("养成", "养成树最大等级", "TREE_MAX", 10, "8 – 15", "每分支等级上限。")

# ============================================================
# 02 合成与充能
# ============================================================
ws = wb.create_sheet("02_合成与充能")
ws["A1"] = "方块值 → 充能 / 结晶 / 各货币 / 附魔伤害 换算表"
ws["A1"].font = TITLE_FONT
head(ws, 3, ["方块值 v", "log₂v", "充能值", "是否超载", "品质", "威力系数",
             "附魔伤害(0★EP)", "CV(合成)", "CV(升华)", "星核", "核心", "金币(关1)", "碎片(关1)"],
     [11, 8, 10, 10, 8, 10, 15, 11, 11, 9, 8, 11, 11])

OVER = {256: (1, 2.0), 512: (2, 3.5), 1024: (3, 6.0), 2048: (4, 10.0), 4096: (5, 16.0)}
r = 4
first_row = r
for i in range(2, 14):  # 4 ... 8192
    v = 2 ** i
    ws.cell(row=r, column=1, value=v).border = BOX
    ws.cell(row=r, column=2, value=i).border = BOX
    c = ws.cell(row=r, column=3, value=f"={CONST['CHARGE_K']}*(B{r}-1)"); c.fill = CALC_FILL; c.border = BOX
    star, mult = OVER.get(v, (0, ""))
    ws.cell(row=r, column=4, value=("超载" if v in OVER else "—")).border = BOX
    ws.cell(row=r, column=5, value=(f"{star}★" if v in OVER else "—")).border = BOX
    if v in OVER:
        ws.cell(row=r, column=6, value=mult).border = BOX
        c = ws.cell(row=r, column=7, value=f"={CONST['EP_BASE']}*F{r}") ; c.fill = CALC_FILL; c.border = BOX
    else:
        ws.cell(row=r, column=6, value="—").border = BOX
        ws.cell(row=r, column=7, value="—").border = BOX
    c = ws.cell(row=r, column=8, value=f"=A{r}"); c.fill = CALC_FILL; c.border = BOX
    c = ws.cell(row=r, column=9, value=f"=IF(A{r}>={CONST['SUB_TH']},A{r}*{CONST['SUB_MULT']},0)"); c.fill = CALC_FILL; c.border = BOX
    c = ws.cell(row=r, column=10, value=f"=H{r}*{CONST['K_STAR']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    core = 8 if v >= 2048 else 3 if v >= 512 else 1 if v >= 128 else 0
    ws.cell(row=r, column=11, value=core).border = BOX
    c = ws.cell(row=r, column=12, value=f"=H{r}*{CONST['K_GOLD']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=13, value=f"=H{r}*{CONST['K_SHARD']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    r += 1
last_row = r - 1

r += 1
r = sec(ws, r, "共鸣倍率（连续同元素触发）")
head(ws, r, ["连续次数", "威力加成", "等效附魔伤害", "说明"], [12, 12, 15, 70])
r += 1
ws.cell(row=r, column=1, value=1).border = BOX
ws.cell(row=r, column=2, value=1.0).border = BOX
c = ws.cell(row=r, column=3, value=f"={CONST['EP_BASE']}*B{r}"); c.fill = CALC_FILL; c.border = BOX
ws.cell(row=r, column=4, value="首次触发，基准。").border = BOX
r += 1
ws.cell(row=r, column=1, value=2).border = BOX
c = ws.cell(row=r, column=2, value=f"={CONST['RES2']}"); c.fill = CALC_FILL; c.border = BOX
c = ws.cell(row=r, column=3, value=f"={CONST['EP_BASE']}*B{r}"); c.fill = CALC_FILL; c.border = BOX
ws.cell(row=r, column=4, value="轮盘排 [X,X] 的收益。").border = BOX
r += 1
ws.cell(row=r, column=1, value="3+").border = BOX
c = ws.cell(row=r, column=2, value=f"={CONST['RES3']}"); c.fill = CALC_FILL; c.border = BOX
c = ws.cell(row=r, column=3, value=f"={CONST['EP_BASE']}*B{r}"); c.fill = CALC_FILL; c.border = BOX
ws.cell(row=r, column=4, value="附带「元素余韵」：持续类效果时长 +50%。轮盘排 [X,X,X] 是爆发构筑核心。").border = BOX
r += 2
r = note(ws, r, "设计校验：一次 2048 超载 = 4★ × EP = 600 伤害 = 15 个第 1 关小兵。这是玩家截图分享的 jackpot 时刻。")
r = note(ws, r, "设计校验：最低合成（4）只给 4 充能，最高（2048）给 40。差距 10 倍 —— 保证大数字的战略价值，同时小合成不至于毫无意义。")
r = note(ws, r, "调参方向：若埋点显示每波小附魔 > 6 次，升高 CHARGE_MAX 或降低 CHARGE_K；若 < 2 次则反向。")

# ============================================================
# 03 波次节奏模型
# ============================================================
ws = wb.create_sheet("03_波次节奏模型")
ws["A1"] = "波次节奏模型 —— 估算每波产出（★ MERGE_RATE 与平均合成值必须用 Monte Carlo 校准）"
ws["A1"].font = TITLE_FONT
head(ws, 3, ["关卡", "生成池", "生成期望", "平均合成值", "每波步数", "合成次数/波",
             "ΣCV/波", "充能/波", "小附魔/波", "星核/波", "金币/波", "碎片/波", "目标校验"],
     [7, 16, 11, 13, 11, 12, 11, 11, 12, 10, 11, 10, 26])

TIERS = [
    (1, "2(90%)/4(10%)", 2.2, 7.0),
    (2, "2(90%)/4(10%)", 2.2, 8.0),
    (3, "4(80%)/8(20%)", 4.8, 16.0),
    (4, "4(80%)/8(20%)", 4.8, 18.0),
    (5, "8(80%)/16(20%)", 9.6, 34.0),
    (6, "8(80%)/16(20%)", 9.6, 38.0),
    (7, "16(85%)/32(15%)", 18.4, 64.0),
    (8, "16(85%)/32(15%)", 18.4, 70.0),
    (9, "32(85%)/64(15%)", 36.8, 130.0),
    (10, "32(85%)/64(15%)", 36.8, 140.0),
    (12, "32(85%)/64(15%)", 36.8, 155.0),
    (15, "32(85%)/64(15%)", 36.8, 170.0),
]
r = 4
for stage, pool, exp_v, avg_merge in TIERS:
    ws.cell(row=r, column=1, value=stage).border = BOX
    ws.cell(row=r, column=2, value=pool).border = BOX
    ws.cell(row=r, column=3, value=exp_v).border = BOX
    c = ws.cell(row=r, column=4, value=avg_merge); c.fill = INPUT_FILL; c.border = BOX
    c = ws.cell(row=r, column=5, value=f"={CONST['WAVE_T']}/{CONST['STEP_REGEN']}+{CONST['STEP_GIFT']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=6, value=f"=E{r}*{CONST['MERGE_RATE']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=7, value=f"=F{r}*D{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=8, value=f"=F{r}*{CONST['CHARGE_K']}*(LOG(D{r},2)-1)"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=9, value=f"=H{r}/{CONST['CHARGE_MAX']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    if stage >= 5:
        c.number_format = "0.00"
        c.font = Font(bold=True, color="C00000")
    c = ws.cell(row=r, column=10, value=f"=G{r}*{CONST['K_STAR']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    c = ws.cell(row=r, column=11, value=f"=G{r}*{CONST['K_GOLD']}*(1+{CONST['RWD_INC']}*(A{r}-1))"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=12, value=f"=G{r}*{CONST['K_SHARD']}*(1+{CONST['RWD_INC']}*(A{r}-1))"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    note_txt = "★目标 3–4 次/波" if stage <= 2 else ("偏高，需校准" if stage >= 7 else "")
    ws.cell(row=r, column=13, value=note_txt).border = BOX
    r += 1

r += 1
r = note(ws, r, "【模型警告】「平均合成值」列是估算值，标记黄色。真实值必须用 10,000 局 Monte Carlo 统计得出，这是纸面原型阶段的首要任务。")
r = note(ws, r, "【设计意图】星核/波 目标 20–30，即每波可释放约 1 次战术指令。若超过 50，说明玩家能靠指令无限续命，战斗压力失效。")
r = note(ws, r, "【设计意图】小附魔/波 随关卡上升是刻意的 —— 后期玩家应该感到「元素轰炸」，这是爽感成长曲线的一部分，但也必须监控不超 6 次。")

# ============================================================
# 04 关卡与敌人
# ============================================================
ws = wb.create_sheet("04_关卡与敌人")
ws["A1"] = "关卡缩放 —— 敌人强度 vs 玩家战力成长（含自然撞墙点判定）"
ws["A1"].font = TITLE_FONT
head(ws, 3, ["关卡 n", "敌人HP倍率", "敌人伤害倍率", "敌人速度倍率", "关卡收益倍率",
             "玩家战力倍率", "余量比", "状态", "第5波敌人总HP(估)", "玩家DPS(估)"],
     [8, 12, 13, 13, 12, 13, 10, 16, 16, 13])

r = 4
for n in range(1, 16):
    ws.cell(row=r, column=1, value=n).border = BOX
    c = ws.cell(row=r, column=2, value=f"={CONST['HP_BASE']}^(A{r}-1)*(1+{CONST['HP_SLIN']}*(A{r}-1))"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=3, value=f"={CONST['DMG_BASE']}^(A{r}-1)"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=4, value=f"=1+{CONST['SPD_SCL']}*(A{r}-1)"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    if n == 1:
        c = ws.cell(row=r, column=5, value=1.0); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
        c = ws.cell(row=r, column=6, value=1.0); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.000"
    else:
        c = ws.cell(row=r, column=5, value=f"=E{r-1}+{CONST['RWD_INC']}+0.1*(A{r}-2)"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
        # 玩家战力：上一关 × (1 + 单卡效果 × 衰减^(n-2))^每关卡数
        c = ws.cell(row=r, column=6, value=f"=F{r-1}*(1+{CONST['CARD_PWR']}*{CONST['CARD_DECAY']}^(A{r}-2))^{CONST['CARD_N']}")
        c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.000"
    c = ws.cell(row=r, column=7, value=f"=F{r}/B{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=8, value=f'=IF(G{r}>1.15,"轻松",IF(G{r}>0.85,"势均力敌",IF(G{r}>0.6,"吃力","撞墙")))')
    c.fill = CALC_FILL; c.border = BOX; c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=r, column=9, value=f"=1400*B{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    c = ws.cell(row=r, column=10, value=f"={CONST['PLANT_DPS']}*{CONST['SLOT_BASE']}*F{r}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    r += 1

r += 1
r = note(ws, r, "【余量比判读】>1.15 轻松 / 0.85–1.15 势均力敌 / 0.6–0.85 吃力 / <0.6 撞墙。")
r = note(ws, r, "★【重要】余量比只是 DPS 上限，是乐观估计。玩家真正的死因不是「打不动」，而是「漏过去的敌人把核心打爆」。必须用下面的漏怪模型判定实际死亡关卡。")
r += 1

r = sec(ws, r, "漏怪预算模型 —— 真实死亡关卡判定")
r = note(ws, r, "逻辑：当 DPS 余量比 < 1 时，缺口比例的敌人会漏过防线；漏怪数 × 敌人伤害 > 核心 HP 即阵亡。")
head(ws, r, ["关卡 n", "余量比(基准)", "可承受漏怪数", "新手(合成率0.5)", "中等(1.2)", "高手(1.8)", "满养成高手"], [8, 13, 14, 17, 14, 14, 15])
r += 1
for n in range(1, 12):
    src_row = n + 3
    ws.cell(row=r, column=1, value=n).border = BOX
    c = ws.cell(row=r, column=2, value=f"=F{src_row}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.00"
    c = ws.cell(row=r, column=3, value=f"={CONST['CORE_HP']}/({CONST['AVG_DMG']}*{CONST['DMG_BASE']}^(A{r}-1))")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    for idx, (mrs, upg) in enumerate([("MR_NOVICE", 1), (None, 1), ("MR_EXPERT", 1), ("MR_EXPERT", 1.6)], start=4):
        if mrs:
            skill = f"(1-{CONST['ENCH_SHARE']}+{CONST['ENCH_SHARE']}*{CONST[mrs]}/{CONST['MERGE_RATE']})"
        else:
            skill = "1"
        f = f'=IF({CONST["ENEMY_TOT"]}*MAX(0,1-B{r}*{skill}*{upg})<C{r},"存活","阵亡")'
        c = ws.cell(row=r, column=idx, value=f); c.fill = CALC_FILL; c.border = BOX
        c.alignment = Alignment(horizontal="center")
    r += 1
r += 1
r = note(ws, r, "【模型输出（当前配置 HP_BASE=1.55）】新手 3 关 / 中等 6 关 / 高手 7 关 / 满养成高手 9 关。这是设计目标曲线，与「平均到达 4–6 关」的埋点目标自洽。")
r = note(ws, r, "【技术水平权重】附魔占输出 45%（ENCH_SHARE）。新手合成率 0.5 → 技术系数 0.74；高手 1.8 → 1.23。这意味着技术差距约 1.65 倍战力，是健康的技巧回报。")
r = note(ws, r, "【调参方向】若新手在 3 关就阵亡导致挫败：降低 HP_BASE 至 1.50，或提高第 1–2 关的波次时长给更多缓冲。若中等玩家超过 7 关：升高 HP_BASE 至 1.60。")
r += 1

r = sec(ws, r, "敌人基准（第 1 关）")
head(ws, r, ["类型", "HP", "速度(格/秒)", "撞核心伤害", "护甲", "掉落材料", "设计目的"], [14, 9, 12, 12, 9, 10, 60])
r += 1
for row in [
    ("小兵 Grunt", 40, 0.35, 5, 0, 0, "基准单位，定义 DPS 需求底线。"),
    ("群聚 Swarm", 12, 0.50, 2, 0, 0, "惩罚单体流，奖励火元素与溅射。"),
    ("迅捷 Swift", 22, 0.75, 3, 0, 0, "惩罚合成太慢，是时间压力的主要载体。"),
    ("重甲 Armor", 140, 0.22, 12, 0.30, 0, "惩罚穿透/破甲不足，奖励雷元素。"),
    ("破坏者 Breaker", 200, 0.30, 15, 0.10, 0, "攻击并摧毁植物格，惩罚固守不动的编队。"),
    ("精英 Elite", 400, 0.28, 25, 0.15, "1–2", "携带元素护盾，检验玩家的轮盘编排。"),
    ("Boss", 1200, 0.20, 40, 0.20, "3–5", "双元素护盾，关卡高潮。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i == 7: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "波次组成（第 1 关）")
head(ws, r, ["波", "组成", "时长(秒)", "敌人总HP", "设计意图"], [6, 34, 11, 12, 60])
r += 1
for row in [
    (1, "6 小兵", 30, 240, "教学波。不可能失败。"),
    (2, "4 小兵 + 4 群聚", 35, 208, "引入群体压力，让玩家理解火/溅射的价值。"),
    (3, "2 重甲 + 6 小兵", 45, 520, "引入护甲，制造第一次「打不动」的挫败。"),
    (4, "6 迅捷 + 2 重甲", 45, 412, "引入时间压力，逼玩家加快合成。"),
    (5, "1 精英(火盾) + 8 小兵 + 4 群聚", 60, 768, "Boss 波。首次检验轮盘编排是否针对护盾。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i == 5: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ============================================================
# 05 植物
# ============================================================
ws = wb.create_sheet("05_植物")
ws["A1"] = "植物数值与成长模型"
ws["A1"].font = TITLE_FONT
r = note(ws, 2, "成长公式：DPS = 基础 × (1 + 每星DPS加成×(星级−1)) × (1 + 每级DPS加成×(等级−1)) × 养成树枝叶加成")
head(ws, 3, ["植物", "元素", "攻击形态", "基础DPS", "射程(格)", "特性", "大招(4核心/25s)"], [13, 8, 13, 10, 10, 44, 46])
r = 4
for row in [
    ("焰心草", "火", "直射单体", 12, 3, "命中附加灼烧：3 秒内每秒 3 伤害", "全场灼烧 8s，每秒 EP×0.15"),
    ("潮汐藤", "水", "环形 AoE", 7, 1.5, "范围内敌人 −20% 移速（持续）", "全体击退 2 格 + 减速 50%（4s）"),
    ("荆棘木", "木", "穿透直线", 11, "无限", "无视前排阻挡，贯穿本行", "贯穿全屏光束，EP×3.0"),
    ("辉光苔", "光", "辅助(无攻击)", 0, "—", "每 10s：+1 步 或 +8 充能（玩家预设）", "立即 +4 步 +60 充能，全植物攻速 +50%（6s）"),
    ("雷鸣花", "雷", "溅射", 10, 2.5, "命中溅射周围 50% 伤害", "15 次随机落雷，每次 EP×0.4"),
    ("霜晶莲", "冰", "直射单体", 8, 3, "命中减速 30%（2s）", "全场冻结 4s，冻结期间受伤 +25%"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 6: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "星级 / 等级成长曲线（以荆棘木 基础DPS=11 为例）")
head(ws, r, ["等级\\星级", "1★", "2★", "3★", "4★", "5★"], [12, 11, 11, 11, 11, 11])
r += 1
for lv in [1, 10, 20, 30, 40, 50]:
    c = ws.cell(row=r, column=1, value=f"Lv.{lv}"); c.border = BOX
    for s in range(1, 6):
        col = get_column_letter(s + 1)
        f = f"=11*(1+{CONST['STAR_MULT']}*({s}-1))*(1+{CONST['LV_MULT']}*({lv}-1))"
        c = ws.cell(row=r, column=s + 1, value=f); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0"
    r += 1
r += 1
r = note(ws, r, "【设计校验】Lv.50 / 5★ 的荆棘木 = 11 × 1.8 × 1.98 = 39.2 DPS，是初始值的 3.56 倍。配合编队位扩展至 6（伙伴位 50% 效率 → ×1.5），植物侧满养成合计约 5.3 倍。")
r = note(ws, r, "★★【平衡红线 · 最重要的一个】养成 vs 局内卡牌的战力占比，建议 60 : 40。")
r = note(ws, r, "   · 若养成占比过高（>75%）：肉鸽三选一沦为过场，局内构筑失去意义，游戏退化成纯数值养成。")
r = note(ws, r, "   · 若养成占比过低（<40%）：H5 产品失去留存与付费主驱动，长期 LTV 崩塌。")
r = note(ws, r, "   · 60:40 的依据：养成是 H5 商业化的主体，但局内卡牌必须保留「每局体验不同」的价值 —— 它的意义在多样性而非数值量级。")
r = note(ws, r, "   · 监控方式：分别统计「满养成 + 零卡牌」与「零养成 + 满卡牌」两种极端配置下的可达关卡，二者之比应落在 1.3 – 1.8 之间。")
r += 1

r = sec(ws, r, "六元素效果表（EP = 附魔强度）")
head(ws, r, ["元素", "定位", "效果（0★ / EP=60）", "设计意图"], [8, 12, 46, 50])
r += 1
for row in [
    ("火", "群体爆发", "全场敌人立即受到 EP×1.0 伤害；已灼烧目标 +30%", "清群首选，最直观的爽感来源。"),
    ("雷", "单体爆发", "连锁闪电弹跳 5 次，每次 EP×0.5，优先最前排", "打 Boss / 重甲。与火形成群-单分工。"),
    ("冰", "控制+增伤", "全场定身 3s，冻结目标受伤 +25%", "救场技，为爆发铺路。进攻向。"),
    ("水", "持续减速", "移速 −40%，持续 5s，造成伤害 −20%", "拖时间，给棋盘争取步数。防守向。"),
    ("木", "己方增益", "植物攻速 +35%，持续 8s；核心回复 EP×0.2", "长线收益，配合高 DPS 编队才有价值。"),
    ("光", "资源调度", "清除棋盘所有 ≤4 方块；+2 步；植物伤害 +25%（8s）", "唯一直接强化棋盘的元素，是解卡保底。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 3: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
r = sec(ws, r, "元素护盾（让轮盘编排有外部约束）")
head(ws, r, ["护盾类型", "效果", "破解方式", "设计目的"], [16, 40, 30, 46])
r += 1
for row in [
    ("元素护盾 · X", "受到非 X 元素伤害 −60%", "用 X 元素附魔 / X 属性植物攻击", "防止玩家把轮盘排成固定最优解。"),
    ("元素易伤 · X", "受到 X 元素伤害 +100%", "针对性编排轮盘", "奖励开局前的配置决策。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ============================================================
# 06 卡牌
# ============================================================
ws = wb.create_sheet("06_卡牌")
ws["A1"] = "卡池清单与稀有度权重"
ws["A1"].font = TITLE_FONT
r = sec(ws, 2, "稀有度权重（按关卡）")
head(ws, 3, ["关卡", "普通", "稀有", "史诗", "传说", "说明"], [8, 10, 10, 10, 10, 60])
r = 4
for row in [
    (1, 0.70, 0.25, 0.05, 0.00, "教学期，几乎只有基础卡。"),
    (2, 0.60, 0.30, 0.09, 0.01, "首次见到传说，制造惊喜。"),
    (3, 0.50, 0.33, 0.15, 0.02, "构筑开始成型。"),
    (4, 0.42, 0.35, 0.18, 0.05, "构筑发力期。"),
    ("5+", 0.35, 0.35, 0.22, 0.08, "后期。同关第 4、5 波再 +5% 稀有度权重。"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if 2 <= i <= 5: c.number_format = "0%"
        if i == 6: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "卡池")
head(ws, r, ["类别", "卡名", "稀有度", "效果", "设计备注"], [11, 16, 9, 50, 46])
r += 1
CARDS = [
    ("棋盘", "稳健", "普通", "步数上限 +1", "最基础的节奏卡，前期必拿。"),
    ("棋盘", "疾风", "普通", "步数回复速度 +15%", "与稳健形成「容量 vs 速率」选择。"),
    ("棋盘", "聚能", "普通", "充能条上限 −10", "等价于附魔频率 +11%。"),
    ("棋盘", "精准", "稀有", "生成值为高档的概率 +25%", "隐性加速，体感不明显但收益高。"),
    ("棋盘", "净化术", "稀有", "立即清除所有 ≤8 方块并 +2 步", "救急卡。"),
    ("棋盘", "丰饶", "史诗", "每次移动额外生成 1 个方块", "★ 双刃卡：产出翻倍，但棋盘更容易满。制造真实抉择。"),
    ("棋盘", "连锁核心", "传说", "一次移动中第 3 次及以后的合并，充能与 CV 翻倍", "★ 奖励布局技巧，是高手与普通玩家的分水岭。"),
    ("战斗", "锋锐", "普通", "全植物伤害 +15%", "基准战力卡。"),
    ("战斗", "连射", "普通", "全植物攻速 +12%", "与锋锐近似等效，制造选择困难。"),
    ("战斗", "远眺", "普通", "全植物射程 +1", "价值依赖编队，情境性强。"),
    ("战斗", "破甲", "稀有", "无视 30% 护甲", "针对重甲/Boss 波。"),
    ("战斗", "暴击", "稀有", "暴击率 +10%，暴击伤害 150%", "期望 +5% 伤害，但方差大，适合赌狗。"),
    ("战斗", "荆棘", "史诗", "敌人撞核心时反伤 200", "被动防守，缓解后期压力。"),
    ("战斗", "共生", "传说", "每个存活植物使全植物伤害 +8%", "配合满编队 = +48%，后期最强战斗卡。"),
    ("附魔", "元素亲和 · X", "普通", "X 元素威力 +30%（六种各一张）", "构筑核心。六张同名系卡。"),
    ("附魔", "充能加速", "普通", "每次合成充能 +25%", "等价于附魔频率 +25%。"),
    ("附魔", "共鸣强化", "稀有", "共鸣加成 1.4→1.6，1.8→2.2", "★ 纯色队的核心卡。"),
    ("附魔", "双生", "史诗", "附魔时同时触发轮盘下一位元素，威力各 60%", "改变轮盘节奏，需要重新规划。"),
    ("附魔", "超载", "史诗", "大附魔品质 +1 级（512 视为 1024）", "★ 大数字流的核心卡。"),
    ("附魔", "星爆", "传说", "2048 附魔威力 +100%，且必定暴击", "极限流冲刺卡。"),
    ("经济", "贪婪", "普通", "金币 +25%", "收益型，适合准备收工时拿。"),
    ("经济", "采集", "普通", "碎片产出 +50%", "长线收益。"),
    ("经济", "结晶", "稀有", "所有 CV × 1.15", "通用收益卡，永远不亏但也不出彩。"),
    ("经济", "升华", "稀有", "自动升华阈值降至 256", "★ 兑现流核心卡，与「贪婪核心」互斥。"),
    ("经济", "囤积", "史诗", "大招核心消耗 −1，星核产出 +20%", "双重收益，史诗级定价合理。"),
    ("战术", "战术家", "普通", "星核产出 +30%", "提升救急能力。"),
    ("战术", "时间膨胀", "稀有", "每波开始 +3 步", "等价于每波 +15% 步数。"),
    ("战术", "急救", "稀有", "核心 HP 首次低于 30% 时回复 40%", "一次性保险。"),
    ("战术", "备用核心", "史诗", "核心归零时以 30% HP 复活一次", "★ 直接改变继续/收工的期望值计算。"),
    ("战术", "贪婪核心", "传说", "关闭自动升华", "★ 高风险高回报的极限流开关，是升华机制的策略深度所在。"),
]
for row in CARDS:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 4: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
r += 1
r = note(ws, r, "卡池当前 30 张（含 6 张元素亲和变体则共 35 张）。V2 完整版目标 ≥ 60 张，否则 4–6 关后卡池枯竭，构筑深度崩塌。")
r = note(ws, r, "平衡红线：同名卡重复时效果加法叠加而非乘法，防止指数爆炸。若埋点显示某卡选取率 > 70%，说明该卡超模或同类卡过少。")
r = note(ws, r, "建议补充机制：卡牌升阶 —— 同名卡可合并强化，缓解后期卡池枯竭（见 GDD 风险 R5）。")

# ============================================================
# 07 风险决策 EV
# ============================================================
ws = wb.create_sheet("07_风险决策EV")
ws["A1"] = "继续 / 收工 期望值计算器"
ws["A1"].font = TITLE_FONT
r = note(ws, 2, "公式：收工 EV = P ｜ 继续 EV = p×(P+R) + (1−p)×b×P ｜ 继续优于收工 ⟺ p > (1−b)P / ((1−b)P + R)")
r = 3
r = sec(ws, r, "主计算")
head(ws, r, ["参数", "值", "说明"], [30, 14, 70])
r += 1
P_CELL = None
rows = [
    ("当前累积收益池 P", 1000, "INPUT_FILL", "任意单位。EV 结论与 P 的绝对值无关，只与 R/P 比例有关。"),
    ("下一关预期收益 R", None, "CALC", f"= 当前池 P × R_RATIO（见 01 表）"),
    ("失败保底比例 b", None, "CALC", "引用 01 表 B_FAIL"),
    ("玩家自估通关概率 p", 0.60, "INPUT_FILL", "玩家的主观判断。设计目标是让这个阈值足够低。"),
]
ws.cell(row=r, column=1, value="当前累积收益池 P").border = BOX
c = ws.cell(row=r, column=2, value=1000); c.fill = INPUT_FILL; c.border = BOX
ws.cell(row=r, column=3, value="任意单位。EV 结论与 P 的绝对值无关，只与 R/P 比例有关。").border = BOX
P_ref = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="下一关预期收益 R").border = BOX
c = ws.cell(row=r, column=2, value=f"={P_ref}*{CONST['R_RATIO']}"); c.fill = CALC_FILL; c.border = BOX
ws.cell(row=r, column=3, value="= P × R_RATIO。R_RATIO 由关卡收益倍率递增推导。").border = BOX
R_ref = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="失败保底比例 b").border = BOX
c = ws.cell(row=r, column=2, value=f"={CONST['B_FAIL']}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0%"
ws.cell(row=r, column=3, value="引用 01 核心常量 B_FAIL。").border = BOX
b_ref = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="玩家自估通关概率 p").border = BOX
c = ws.cell(row=r, column=2, value=0.60); c.fill = INPUT_FILL; c.border = BOX; c.number_format = "0%"
ws.cell(row=r, column=3, value="玩家主观判断。设计目标是让临界阈值足够低，从而鼓励冒险。").border = BOX
p_ref = f"$B${r}"; r += 1
r += 1

head(ws, r, ["指标", "值", "说明"], [30, 14, 70])
r += 1
ws.cell(row=r, column=1, value="收工 EV").border = BOX
c = ws.cell(row=r, column=2, value=f"={P_ref}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
ws.cell(row=r, column=3, value="确定性收益。").border = BOX
ev_stop = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="继续 EV").border = BOX
c = ws.cell(row=r, column=2, value=f"={p_ref}*({P_ref}+{R_ref})+(1-{p_ref})*{b_ref}*{P_ref}")
c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
ws.cell(row=r, column=3, value="= p×(P+R) + (1−p)×b×P").border = BOX
ev_go = f"$B${r}"; r += 1

ws.cell(row=r, column=1, value="EV 差值（继续 − 收工）").border = BOX
c = ws.cell(row=r, column=2, value=f"={ev_go}-{ev_stop}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
c.font = Font(bold=True)
ws.cell(row=r, column=3, value="> 0 表示理性上应该选择继续。").border = BOX
r += 1

ws.cell(row=r, column=1, value="临界通关概率 p*").border = BOX
c = ws.cell(row=r, column=2, value=f"=(1-{b_ref})*{P_ref}/((1-{b_ref})*{P_ref}+{R_ref})")
c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0%"
c.font = Font(bold=True, color="C00000")
ws.cell(row=r, column=3, value="★ p > p* 时继续是正期望。这是整个风险系统的核心数字。").border = BOX
r += 1
r += 1

r = sec(ws, r, "保底比例 b 的敏感性分析（★ 冒险倾向主旋钮）")
head(ws, r, ["保底 b", "临界概率 p*", "玩家行为预测", "适用产品目标"], [11, 14, 46, 46])
r += 1
SENS_START = r
for b in [0.0, 0.2, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0]:
    c = ws.cell(row=r, column=1, value=b); c.border = BOX; c.number_format = "0%"
    c = ws.cell(row=r, column=2, value=f"=(1-A{r})/((1-A{r})+{CONST['R_RATIO']})")
    c.fill = CALC_FILL; c.border = BOX; c.number_format = "0.0%"
    r += 1
PRED = {
    0.0: ("极度保守，玩家会提前收工", "硬核产品，追求策略纯度"),
    0.2: ("偏保守，多数玩家 3–4 关收工", "中核产品，单局时长可控"),
    0.3: ("略保守", "—"),
    0.4: ("★ 平衡点，决策有张力", "★ 推荐：H5 休闲偏中核"),
    0.5: ("略激进", "—"),
    0.6: ("激进，玩家几乎不主动收工", "追求高重开率与广告曝光"),
    0.8: ("极度激进，收工按钮形同虚设", "纯广告变现产品"),
    1.0: ("无风险，继续永远最优", "决策消失，不推荐"),
}
rr = SENS_START
for b in [0.0, 0.2, 0.3, 0.4, 0.5, 0.6, 0.8, 1.0]:
    pred, goal = PRED[b]
    c = ws.cell(row=rr, column=3, value=pred); c.border = BOX; c.alignment = Alignment(wrap_text=True, vertical="top")
    c = ws.cell(row=rr, column=4, value=goal); c.border = BOX; c.alignment = Alignment(wrap_text=True, vertical="top")
    if b == 0.4:
        for col in (1, 2, 3, 4):
            ws.cell(row=rr, column=col).fill = PatternFill("solid", fgColor="FFF2CC")
    rr += 1
r = rr + 1
r = note(ws, r, "【关键洞察】保底比例 b 从 0.2 提到 0.6，临界概率从 50% 降到 33% —— 玩家会明显更爱冒险。这是调节「重开率」最廉价的旋钮，优先用它做 A/B 测。")
r = note(ws, r, "★【隐藏的美妙性质】R/P 比例随关卡自然变化：第 2 关约 1.55（p*≈28%，极度鼓励冒险）→ 第 6 关约 0.53（p*≈53%，必须谨慎）。也就是说游戏在早期自动怂恿玩家冒进、后期自动转为审慎 —— 这是收益倍率递增结构带来的免费礼物，不要在调参时破坏它。")
r = note(ws, r, "【注意】卡牌「备用核心」（核心归零时以 30% HP 复活一次）等价于把 b 临时抬高，会显著改变玩家在持有该卡时的决策。这张卡的实际价值远高于其描述。")
r = note(ws, r, "【UI 强制要求】关卡结算界面必须同屏显示：当前池 / 下一关倍率与威胁等级 / 继续的预估收益 +R / 失败保底规则。缺一项，这个决策就不成立。")

# ============================================================
# 08 养成与花园
# ============================================================
ws = wb.create_sheet("08_养成与花园")
ws["A1"] = "星尘养成树与花园产出模型"
ws["A1"].font = TITLE_FONT
r = sec(ws, 2, "养成树成本模型")
r = note(ws, r, "成本公式：Cost(n) = 基础成本 × 增长系数^(n−1)")
head(ws, r, ["等级", "单级成本(星尘)", "累计成本", "根系效果", "枝叶效果", "花蕾效果", "果实效果"], [8, 15, 13, 26, 26, 22, 32])
r += 1
for lv in range(1, 11):
    ws.cell(row=r, column=1, value=lv).border = BOX
    c = ws.cell(row=r, column=2, value=f"=ROUND({CONST['TREE_BASE']}*{CONST['TREE_GROW']}^(A{r}-1),0)"); c.fill = CALC_FILL; c.border = BOX
    c = ws.cell(row=r, column=3, value=(f"=B{r}" if lv == 1 else f"=C{r-1}+B{r}")); c.fill = CALC_FILL; c.border = BOX
    ws.cell(row=r, column=4, value=f"步数+0.2 / 回复+3% / 充能条−2%").border = BOX
    ws.cell(row=r, column=5, value=f"植物伤害+3% / 核心HP+4%").border = BOX
    ws.cell(row=r, column=6, value=f"全元素威力+4%").border = BOX
    ws.cell(row=r, column=7, value=f"金币+4% / 碎片+6% / 花园产出+5%").border = BOX
    r += 1
ws.cell(row=r, column=1, value="满级合计").font = Font(bold=True)
c = ws.cell(row=r, column=3, value=f"=SUM(B{r-10}:B{r-1})*4"); c.fill = CALC_FILL; c.border = BOX; c.font = Font(bold=True)
ws.cell(row=r, column=4, value="4 分支 × 单分支累计").font = Font(bold=True)
r += 2
r = note(ws, r, "【当前配置】基础 40 / 增长 1.30 → 单分支满级约 1,705 星尘，全树约 6,820 星尘。早期花园产出 200–500/天，后期 2,000+/天 → 约 3–5 周长线目标。")
r = note(ws, r, "【调参方向】H5 休闲产品建议长线不超过 4 周。若超长，把增长系数降到 1.25（全树约 5,300）；若希望更长线支撑赛季，升到 1.35。")
r += 1

r = sec(ws, r, "花园产出模型")
head(ws, r, ["稀有度", "星尘/小时(基)", "30min", "2h", "6h", "6 盆全传说/6h"], [12, 14, 10, 10, 10, 18])
r += 1
for name, base in [("普通", 8), ("稀有", 15), ("史诗", 28), ("传说", 50)]:
    ws.cell(row=r, column=1, value=name).border = BOX
    c = ws.cell(row=r, column=2, value=base); c.fill = INPUT_FILL; c.border = BOX
    for col, hrs in [(3, 0.5), (4, 2), (5, 6)]:
        c = ws.cell(row=r, column=col, value=f"=B{r}*{hrs}"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "0"
    c = ws.cell(row=r, column=6, value=f"=B{r}*6*6"); c.fill = CALC_FILL; c.border = BOX; c.number_format = "#,##0"
    r += 1
r += 1
r = note(ws, r, "产出速率 = 稀有度基础 × (1 + 星级×0.1) × (1 + 养成树果实分支等级×0.05)")
r = note(ws, r, "离线计时上限 12 小时（防纯挂机，同时保证次日回归有收益 → 这是次留的核心钩子）。")
r = note(ws, r, "花盆：初始 2，最多 6。金币解锁成本 5,000 / 20,000 / 60,000 / 150,000。")

# ============================================================
# 09 验收清单
# ============================================================
ws = wb.create_sheet("09_验收清单")
ws["A1"] = "平衡目标与 Broken 判定标准 —— playtest 前必读"
ws["A1"].font = TITLE_FONT
r = note(ws, 2, "原则：先定义「什么算坏」，才能在 playtest 中认出它。所有阈值在此表预先约定，不做事后解释。")
r = 4
r = sec(ws, r, "健康指标（埋点目标）")
head(ws, r, ["类别", "指标", "目标区间", "低于下限的处置", "高于上限的处置"], [10, 32, 18, 40, 40])
r += 1
for row in [
    ("节奏", "每波步数消耗", "25 – 30 步", "步数回复过快或玩家不积极 → 降低回复速率", "步数不够用 → 提高回复速率或降低消耗"),
    ("节奏", "每波合成次数", "30 – 40 次", "棋盘太挤或生成值过低 → 提升生成分层", "棋盘过于宽松 → 提升难度"),
    ("节奏", "每波小附魔触发", "3 – 4 次", "充能条过高 → 降低 CHARGE_MAX", "充能刷屏 → 升高 CHARGE_MAX（>6 必须处理）"),
    ("节奏", "每波战术指令使用", "0.8 – 1.5 次", "星核产出不足 → 提高 K_STAR", "指令续命 → 降低 K_STAR（>2 必须处理）"),
    ("难度", "平均到达关卡", "4 – 6 关", "难度过高 → 降低 HP_BASE 至 1.50", "难度过低 → 升高 HP_BASE 至 1.60"),
    ("难度", "第 1 关失败率", "< 5%", "—（达标即可）", "教学波设计失败，需重做波次组成"),
    ("难度", "无养成新手到达", "3 关", "新手挫败 → 前 2 关难度下调", "元游戏无意义 → 前期难度上调"),
    ("难度", "满养成高手到达", "9 – 11 关", "长线目标不足 → 提高养成上限", "养成碾压 → 压缩植物星级/等级成长"),
    ("难度", "收工关卡众数", "第 4 – 6 关", "玩家不敢冒险 → 提高保底 b", "玩家不敢收工 → 降低保底 b"),
    ("留存", "单局时长", "12 – 20 分钟", "内容不足 → 增加波数或波次时长", "疲劳 → 缩短波次时长至 30s"),
    ("健康", "棋盘卡死触发率", "< 3%", "—（达标即可）", "棋盘太小或生成算法有问题，需修伪随机"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 3: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "Broken 判定（出现任一症状即为设计失败，需回头改系统而非改数值）")
head(ws, r, ["症状", "判定阈值", "根因假设", "处置"], [30, 20, 42, 42])
r += 1
for row in [
    ("战斗毫无压力", "核心 HP 全损率 < 10%", "敌人推进速度或数量不足", "提高速度缩放 SPD_SCL，或降低植物基础 DPS"),
    ("棋盘被忽略", "步数常年满仓（余量 > 60%）", "合成收益不足以驱动玩家操作", "提高 CV 转化系数，或降低步数回复速率"),
    ("附魔刷屏", "每波小附魔 > 6 次", "充能条过低", "升高 CHARGE_MAX 至 130–150"),
    ("卡牌无选择", "某张卡选取率 > 70%", "该卡超模或同类卡过少", "削弱该卡，或补充 2–3 张同类竞品卡"),
    ("经济通胀", "金币日增量 / 日消耗 > 3", "金币汇不足", "增设金币汇（新商店项），或降低 K_GOLD"),
    ("元素单一最优解", "某元素使用率 > 50%", "该元素超模，或克制设计未生效", "调整该元素数值，或增加对应护盾的精英出现率"),
    ("后期棋盘饱和", "第 8 关后棋盘空格 < 3", "升华阈值过高", "降低 SUB_TH 至 256，或提高「贪婪核心」出现率"),
    ("养成碾压局内", "满养成 vs 零养成可达关卡比 > 1.8", "植物星级/等级/编队位成长过大", "压缩 STAR_MULT / LV_MULT / SLOT_EFF"),
    ("局内卡牌无意义", "满卡牌 vs 零卡牌可达关卡比 < 1.3", "卡牌效果或数量不足", "提高 CARD_PWR 或扩充卡池至 60 张"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 3: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 1
r = sec(ws, r, "平衡流程（不可跳过）")
head(ws, r, ["阶段", "活动", "产出", "准入条件"], [18, 46, 34, 34])
r += 1
for row in [
    ("1. 纸面模拟", "Monte Carlo 跑 10,000 局 2048，统计 150 步内的最高方块分布、合成次数分布", "校准 MERGE_RATE 与平均合成值", "写代码之前完成"),
    ("2. 数值原型", "只做棋盘 + 充能 + 附魔伤害的裸 Demo（无战斗）", "验证乐趣假设：精准四连合成触发超载清屏的瞬间是否成立", "纸面模拟完成"),
    ("3. 可玩原型", "加入战斗与波次", "验证压力曲线与注意力分配", "乐趣假设成立"),
    ("4. 小规模测试", "20–50 人试玩，只观察 Feel 问题，不调数值", "Feel 问题清单", "核心循环可玩"),
    ("5. 数值调优", "基于埋点做回归分析", "v0.2 数值", "埋点体系上线"),
    ("6. 软启动 A/B", "测三个主旋钮：STEP_REGEN / B_FAIL / HP_BASE", "上线数值配置", "数值基本稳定"),
]:
    for i, v in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        if i >= 2: c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

r += 2
r = sec(ws, r, "给开发团队的一句话")
c = ws.cell(row=r, column=1, value="V1 垂直切片是生死线。如果 V1（1 主植物 / 3 波 / 2 元素 / 10 张卡）玩起来不爽，加再多内容和养成也救不回来 —— 请在 V1 投入充分的试玩时间，而不是赶着堆内容。")
c.font = Font(bold=True, size=11, color="C00000")
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.row_dimensions[r].height = 40

# 冻结首行
for s in wb.worksheets:
    s.freeze_panes = "A4" if s.title != "00_说明" else "A3"

out = r"C:\Users\creat\WorkBuddy AI\2026-08-31-15-18-31\Balance_星序防线_v0.1.xlsx"
wb.save(out)
print("SAVED:", out)
